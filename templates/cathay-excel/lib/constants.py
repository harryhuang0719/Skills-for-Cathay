"""
Cathay Capital PE Financial Model — Brand Constants & Shared Definitions.

All constants used across the Excel model generator live here:
sheet names, column layout, brand colors, fonts, fills, borders, number formats,
alignment presets, and segment defaults.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. Sheet Names (ordered)
# =============================================================================

SHEET_NAMES = [
    'Cover',
    'Key Assumptions',
    'Revenue Build-up',
    'COGS & OpEx',
    'Income Statement',
    'Balance Sheet',
    'Cash Flow Statement',
    'Working Capital',
    'Debt & CapEx',
    'Returns & Sensitivity',
    'DCF Valuation',
    'Comps',
    'Dashboard',
]
NUM_SHEETS = 13

# =============================================================================
# 2. Column Layout
# =============================================================================

# Column indices (1-based for openpyxl)
COL_LABEL = 1           # A: Row labels
COL_UNIT = 2            # B: Units
COL_NOTE = 3            # C: Notes/references
COL_HIST_START = 4      # D: First historical year (2021)
COL_HIST_END = 7        # G: Last historical year (2024)
COL_FORECAST_START = 8  # H: First forecast year (2025E)
COL_FORECAST_END = 12   # L: Last forecast year (2029E)
COL_LAST = 12

HIST_YEARS = [2021, 2022, 2023, 2024]
FORECAST_YEARS = [2025, 2026, 2027, 2028, 2029]
ALL_YEARS = HIST_YEARS + FORECAST_YEARS
NUM_HIST = len(HIST_YEARS)
NUM_FORECAST = len(FORECAST_YEARS)


# Column letter helper
def col_letter(col_idx):
    """Return Excel column letter for a 1-based index."""
    return get_column_letter(col_idx)


# =============================================================================
# 3. Column Widths
# =============================================================================

COL_WIDTHS = {
    1: 35,   # A: labels
    2: 8,    # B: units
    3: 15,   # C: notes
}
# D-L: data columns, width 14
for _c in range(4, 13):
    COL_WIDTHS[_c] = 14

# =============================================================================
# 4. Brand Colors (Cathay palette, matching PPT skill)
# =============================================================================

CATHAY_RED = '800000'
CATHAY_DARK_RED = '5E0000'
CATHAY_GOLD = 'E8B012'
CATHAY_LIGHT_GOLD = 'F6DC92'
CATHAY_BLACK = '1A1A1A'
CATHAY_DARK_GREY = '595959'
CATHAY_GREY = '808080'
CATHAY_LIGHT_BG = 'F2F2F2'
CATHAY_BLUE = '0070C0'         # hardcoded input cells
CATHAY_GREEN = '00B050'        # cross-sheet links
CATHAY_CHECK_GREEN = '00B050'
CATHAY_CHECK_RED = 'FF0000'

# Pre-built fills
HEADER_FILL = PatternFill('solid', fgColor=CATHAY_RED)
SUBHEAD_FILL = PatternFill('solid', fgColor='FBE9E8')  # light pink
FORECAST_HEADER_FILL = PatternFill('solid', fgColor=CATHAY_GOLD)
ALT_ROW_FILL = PatternFill('solid', fgColor=CATHAY_LIGHT_BG)
CHECK_PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
CHECK_FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')

# Pre-built fonts
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUBHEAD_FONT = Font(name='Calibri', size=10, bold=True, color=CATHAY_RED)
LABEL_FONT = Font(name='Calibri', size=10, color=CATHAY_BLACK)
LABEL_BOLD_FONT = Font(name='Calibri', size=10, bold=True, color=CATHAY_BLACK)
INPUT_FONT = Font(name='Calibri', size=10, color=CATHAY_BLUE)        # blue = hardcoded
FORMULA_FONT = Font(name='Calibri', size=10, color=CATHAY_BLACK)     # black = formula
LINK_FONT = Font(name='Calibri', size=10, color=CATHAY_GREEN)        # green = cross-sheet
UNIT_FONT = Font(name='Calibri', size=9, color=CATHAY_GREY)
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color=CATHAY_GREY)
PCT_FONT = Font(name='Calibri', size=10, color=CATHAY_DARK_GREY)

# Pre-built borders
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
THICK_RIGHT_BORDER = Border(
    right=Side(style='medium', color=CATHAY_BLACK),
)
BOTTOM_TOTAL_BORDER = Border(
    top=Side(style='thin', color=CATHAY_BLACK),
    bottom=Side(style='double', color=CATHAY_BLACK),
)
BOTTOM_SUBTOTAL_BORDER = Border(
    bottom=Side(style='thin', color=CATHAY_BLACK),
)

# =============================================================================
# 5. Number Formats
# =============================================================================

NUM_FMT_NUMBER = '#,##0'
NUM_FMT_DECIMAL = '#,##0.0'
NUM_FMT_PCT = '0.0%'
NUM_FMT_PCT2 = '0.00%'
NUM_FMT_MULTIPLE = '0.0x'
NUM_FMT_CURRENCY_RMB = '¥#,##0'
NUM_FMT_CURRENCY_USD = '$#,##0'
NUM_FMT_NEGATIVE_RED = '#,##0;[Red](#,##0)'
NUM_FMT_NEGATIVE_PAREN = '#,##0_);(#,##0)'
NUM_FMT_YEAR = '0'

# =============================================================================
# 6. Segment Constants
# =============================================================================

NUM_SEGMENTS = 3
SEGMENT_LABELS = ['Segment A', 'Segment B', 'Segment C']

# =============================================================================
# 7. Alignment Presets
# =============================================================================

ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
