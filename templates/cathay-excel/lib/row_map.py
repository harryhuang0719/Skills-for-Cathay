"""
Cathay Capital PE Financial Model — Declarative Row Mapping.

THE MOST CRITICAL FILE IN THE SYSTEM.
Every formula in the model references this map by semantic name, NEVER by raw row number.
This prevents the row-offset bugs that corrupted the previous template.

Usage:
    from row_map import ROWS, SHEETS, row, cell_ref, sheet_cell_ref, data_range

Design Principles:
    1. Every sheet has a dict mapping semantic names to row numbers (1-based, openpyxl convention)
    2. Every row that participates in a formula MUST be in the map
    3. Helper functions look up rows by (sheet_key, item_key)
    4. NO raw row numbers anywhere else in the codebase
"""

from openpyxl.utils import get_column_letter

# =============================================================================
# SHEETS — metadata for each sheet (key → name, index)
# =============================================================================

SHEETS = {
    'cover': {'name': 'Cover', 'index': 1},
    'assumptions': {'name': 'Key Assumptions', 'index': 2},
    'revenue': {'name': 'Revenue Build-up', 'index': 3},
    'cogs_opex': {'name': 'COGS & OpEx', 'index': 4},
    'income_statement': {'name': 'Income Statement', 'index': 5},
    'balance_sheet': {'name': 'Balance Sheet', 'index': 6},
    'cash_flow': {'name': 'Cash Flow Statement', 'index': 7},
    'working_capital': {'name': 'Working Capital', 'index': 8},
    'debt_capex': {'name': 'Debt & CapEx', 'index': 9},
    'returns': {'name': 'Returns & Sensitivity', 'index': 10},
    'dcf': {'name': 'DCF Valuation', 'index': 11},
    'comps': {'name': 'Comps', 'index': 12},
    'dashboard': {'name': 'Dashboard', 'index': 13},
}

# =============================================================================
# ROWS — the master row map
# =============================================================================

ROWS = {

    # =========================================================================
    # Cover (Sheet 1)
    # =========================================================================
    'cover': {
        'company_name': 4,
        'industry': 5,
        'date': 6,
        'analyst': 7,
        'fx_rate': 8,
        'base_currency': 9,
    },

    # =========================================================================
    # Key Assumptions (Sheet 2, ~100 rows)
    # =========================================================================
    'assumptions': {
        'header': 1,
        'year_row': 3,

        # Scenario Toggle
        'scenario_label': 5,
        'scenario_toggle': 6,       # dropdown cell
        'revenue_multiplier': 7,    # =IF logic
        'margin_adjustment': 8,

        # Segment A Revenue Drivers (rows 10-16)
        'seg_a_header': 10,
        'seg_a_volume': 11,
        'seg_a_price': 12,
        'seg_a_utilization': 13,
        'seg_a_revenue': 14,        # =volume x price x utilization x multiplier
        'seg_a_growth': 15,
        'seg_a_pct': 16,            # % of total revenue

        # Segment B Revenue Drivers (rows 18-24)
        'seg_b_header': 18,
        'seg_b_volume': 19,
        'seg_b_price': 20,
        'seg_b_utilization': 21,
        'seg_b_revenue': 22,
        'seg_b_growth': 23,
        'seg_b_pct': 24,

        # Segment C Revenue Drivers (rows 26-32)
        'seg_c_header': 26,
        'seg_c_volume': 27,
        'seg_c_price': 28,
        'seg_c_utilization': 29,
        'seg_c_revenue': 30,
        'seg_c_growth': 31,
        'seg_c_pct': 32,

        # Total Revenue (row 34)
        'total_revenue': 34,        # =seg_a + seg_b + seg_c
        'total_growth': 35,

        # Cost Assumptions (rows 37-42)
        'cost_header': 37,
        'cogs_pct_a': 38,           # COGS as % of segment revenue
        'cogs_pct_b': 39,
        'cogs_pct_c': 40,
        'blended_cogs_pct': 41,
        'gross_margin': 42,         # = 1 - blended_cogs_pct

        # OpEx Assumptions (rows 44-50)
        'opex_header': 44,
        'personnel_pct': 45,
        'rent_pct': 46,
        'marketing_pct': 47,
        'rd_pct': 48,
        'other_opex_pct': 49,
        'total_sga_pct': 50,

        # CapEx & D&A (rows 52-54)
        'capex_header': 52,
        'capex_pct': 53,
        'da_rate': 54,              # D&A as % of prior PP&E

        # Working Capital Days (rows 58-61)
        'wc_header': 58,
        'ar_days': 59,
        'ap_days': 60,
        'inventory_days': 61,

        # Tax & Other (rows 65-68)
        'tax_header': 65,
        'tax_rate': 66,
        'dividend_payout': 67,
        'interest_rate': 68,
    },

    # =========================================================================
    # Revenue Build-up (Sheet 3, ~40 rows)
    # =========================================================================
    'revenue': {
        'header': 1,
        'year_row': 3,

        # Bottom-up
        'bu_header': 5,
        'bu_seg_a': 6,
        'bu_seg_b': 7,
        'bu_seg_c': 8,
        'bu_total': 10,

        # Top-down
        'td_header': 12,
        'tam': 13,
        'penetration': 14,
        'market_share': 15,
        'td_total': 17,             # = tam x penetration x share

        # Reconciliation
        'recon_header': 19,
        'recon_diff': 20,           # = (BU - TD) / TD
        'recon_flag': 21,           # conditional: >10% shows "CHECK"

        # Revenue Mix
        'mix_header': 23,
        'mix_seg_a_pct': 24,
        'mix_seg_b_pct': 25,
        'mix_seg_c_pct': 26,

        # YoY Growth
        'growth_header': 28,
        'growth_total': 29,
        'growth_seg_a': 30,
        'growth_seg_b': 31,
        'growth_seg_c': 32,
    },

    # =========================================================================
    # COGS & OpEx (Sheet 4, ~50 rows)
    # =========================================================================
    'cogs_opex': {
        'header': 1,
        'year_row': 3,

        # COGS by Segment
        'cogs_header': 5,
        'cogs_seg_a': 6,            # = seg_a_revenue x cogs_pct_a
        'cogs_seg_b': 7,
        'cogs_seg_c': 8,
        'total_cogs': 10,
        'cogs_pct_rev': 11,

        # Gross Profit
        'gross_profit': 13,         # = revenue - total_cogs
        'gross_margin': 14,

        # SG&A Breakdown
        'sga_header': 16,
        'personnel': 17,            # = revenue x personnel_pct
        'rent': 18,
        'marketing': 19,
        'rd': 20,
        'other_opex': 21,
        'total_sga': 23,
        'sga_pct_rev': 24,

        # D&A
        'da_header': 26,
        'depreciation': 27,         # = prior PP&E x da_rate
        'amortization': 28,
        'total_da': 29,

        # Total OpEx
        'total_opex': 31,           # = total_sga + total_da

        # EBITDA Bridge
        'ebitda_bridge_header': 33,
        'ebitda': 34,               # = gross_profit - total_sga
        'ebitda_margin': 35,
        'ebit': 37,                 # = ebitda - total_da
        'ebit_margin': 38,
    },

    # =========================================================================
    # Income Statement (Sheet 5, ~35 rows)
    # =========================================================================
    'income_statement': {
        'header': 1,
        'year_row': 3,

        'revenue': 5,
        'cogs': 7,
        'gross_profit': 9,
        'gross_margin': 10,

        'sga': 12,
        'other_income': 13,
        'ebitda': 15,
        'ebitda_margin': 16,

        'da': 18,
        'ebit': 20,
        'ebit_margin': 21,

        'interest_expense': 23,
        'other_fin': 24,
        'ebt': 26,

        'tax': 27,
        'effective_tax_rate': 28,
        'net_income': 30,
        'net_margin': 31,

        # Per-share (optional)
        'shares_outstanding': 33,
        'eps': 34,
    },

    # =========================================================================
    # Balance Sheet (Sheet 6, ~45 rows)
    # =========================================================================
    'balance_sheet': {
        'header': 1,
        'year_row': 3,

        # Current Assets
        'ca_header': 5,
        'cash': 6,                  # = prior cash + net CF
        'accounts_receivable': 7,   # = revenue / 365 x AR days
        'inventory': 8,             # = COGS / 365 x inventory days
        'other_current': 9,
        'total_current_assets': 11,

        # Non-Current Assets
        'nca_header': 13,
        'ppe_net': 14,              # = prior PP&E + CapEx - D&A
        'intangibles': 15,
        'other_noncurrent': 16,
        'total_noncurrent_assets': 18,

        'total_assets': 20,

        # Current Liabilities
        'cl_header': 22,
        'accounts_payable': 23,     # = COGS / 365 x AP days
        'st_debt': 24,
        'other_current_liab': 25,
        'total_current_liab': 27,

        # Non-Current Liabilities
        'ncl_header': 29,
        'lt_debt': 30,
        'other_noncurrent_liab': 31,
        'total_noncurrent_liab': 33,

        'total_liabilities': 35,

        # Equity
        'eq_header': 37,
        'paid_in_capital': 38,
        'retained_earnings': 39,    # = prior RE + NI - dividends
        'total_equity': 41,

        'total_le': 43,             # = total_liabilities + total_equity
        'bs_check': 45,             # = total_assets - total_le (MUST = 0)
    },

    # =========================================================================
    # Cash Flow Statement (Sheet 7, ~35 rows)
    # =========================================================================
    'cash_flow': {
        'header': 1,
        'year_row': 3,

        # Operating
        'op_header': 5,
        'net_income': 6,            # = IS net_income (green link)
        'da': 7,                    # = COGS & OpEx total_da (green link)
        'delta_ar': 8,              # = -(AR_t - AR_t-1)
        'delta_inventory': 9,
        'delta_ap': 10,
        'delta_other_wc': 11,
        'total_delta_wc': 12,
        'operating_cf': 14,         # = NI + D&A + delta_WC

        # Investing
        'inv_header': 16,
        'capex': 17,                # = -CapEx from Debt & CapEx (green link)
        'other_investing': 18,
        'investing_cf': 20,

        # Financing
        'fin_header': 22,
        'debt_drawdown': 23,
        'debt_repayment': 24,
        'dividends': 25,
        'equity_issuance': 26,
        'financing_cf': 28,

        # Summary
        'net_cf': 30,               # = operating + investing + financing
        'beginning_cash': 31,       # = prior period ending cash
        'ending_cash': 32,          # = beginning + net
        'cash_tieout': 34,          # = ending_cash - BS cash (MUST = 0)
    },

    # =========================================================================
    # Working Capital (Sheet 8, ~30 rows)
    # =========================================================================
    'working_capital': {
        'header': 1,
        'year_row': 3,

        # Days (inputs)
        'days_header': 5,
        'ar_days': 6,
        'inventory_days': 7,
        'ap_days': 8,

        # Balances
        'bal_header': 10,
        'ar_balance': 11,           # = revenue / 365 x AR days
        'inventory_balance': 12,    # = COGS / 365 x inventory days
        'ap_balance': 13,           # = COGS / 365 x AP days
        'net_wc': 15,               # = AR + inventory - AP

        # Changes
        'change_header': 17,
        'delta_ar': 18,
        'delta_inventory': 19,
        'delta_ap': 20,
        'delta_net_wc': 22,         # = net_wc_t - net_wc_t-1
    },

    # =========================================================================
    # Debt & CapEx (Sheet 9, ~50 rows)
    # =========================================================================
    'debt_capex': {
        'header': 1,
        'year_row': 3,

        # Debt Tranche 1
        'd1_header': 5,
        'd1_beginning': 6,
        'd1_drawdown': 7,
        'd1_repayment': 8,
        'd1_ending': 9,
        'd1_rate': 10,
        'd1_interest': 11,

        # Debt Tranche 2
        'd2_header': 13,
        'd2_beginning': 14,
        'd2_drawdown': 15,
        'd2_repayment': 16,
        'd2_ending': 17,
        'd2_rate': 18,
        'd2_interest': 19,

        # Debt Summary
        'debt_summary_header': 21,
        'total_debt': 22,           # = d1_ending + d2_ending
        'total_interest': 23,       # = d1_interest + d2_interest
        'total_drawdown': 24,
        'total_repayment': 25,

        # CapEx
        'capex_header': 27,
        'capex_maintenance': 28,
        'capex_growth': 29,
        'total_capex': 30,

        # D&A Schedule
        'da_header': 32,
        'beginning_ppe': 33,
        'plus_capex': 34,
        'less_da': 35,
        'ending_ppe': 36,           # = beginning + capex - DA
    },

    # =========================================================================
    # Returns & Sensitivity (Sheet 10, ~55 rows)
    # =========================================================================
    'returns': {
        'header': 1,
        'year_row': 3,

        # Entry Assumptions
        'entry_header': 5,
        'entry_revenue': 6,
        'entry_net_income': 7,
        'entry_ps_multiple': 8,
        'entry_pe_multiple': 9,
        'entry_ev': 10,
        'entry_net_debt': 11,
        'entry_equity': 12,

        # Exit — P/S Method
        'ps_exit_header': 14,
        'ps_exit_revenue': 15,
        'ps_exit_multiple': 16,
        'ps_exit_ev': 17,
        'ps_exit_net_debt': 18,
        'ps_exit_equity': 19,
        'ps_moic': 20,
        'ps_irr': 21,

        # Exit — P/E Method
        'pe_exit_header': 23,
        'pe_exit_ni': 24,
        'pe_exit_multiple': 25,
        'pe_exit_equity': 26,
        'pe_moic': 27,
        'pe_irr': 28,

        # Sensitivity Table 1: IRR vs Entry x Exit P/E
        'sens1_header': 31,
        'sens1_top_left': 32,       # anchor cell for the sensitivity grid

        # Sensitivity Table 2: MOIC vs Revenue CAGR x Exit Year
        'sens2_header': 43,
        'sens2_top_left': 44,
    },

    # =========================================================================
    # DCF Valuation (Sheet 11, ~40 rows)
    # =========================================================================
    'dcf': {
        'header': 1,
        'year_row': 3,

        # WACC
        'wacc_header': 5,
        'risk_free': 6,
        'beta': 7,
        'mrp': 8,
        'cost_of_equity': 9,        # = Rf + Beta x MRP
        'cost_of_debt': 10,
        'tax_rate': 11,
        'equity_weight': 12,
        'debt_weight': 13,
        'wacc': 14,                 # = E/(D+E) x Ke + D/(D+E) x Kd x (1-t)

        # UFCF
        'ufcf_header': 16,
        'ebit': 17,                 # green link from IS
        'tax_on_ebit': 18,
        'nopat': 19,                # = EBIT x (1-t)
        'plus_da': 20,
        'less_capex': 21,
        'less_delta_wc': 22,
        'ufcf': 24,                 # = NOPAT + D&A - CapEx - delta_WC

        # Discount
        'discount_header': 26,
        'discount_period': 27,      # mid-year: 0.5, 1.5, 2.5, ...
        'discount_factor': 28,      # = 1/(1+WACC)^period
        'pv_fcf': 29,              # = UFCF x discount_factor

        # Terminal Value
        'tv_header': 31,
        'tv_method_toggle': 32,     # "Gordon" or "Exit Multiple"
        'terminal_growth': 33,
        'exit_multiple': 34,
        'terminal_value': 35,       # = IF(Gordon, ..., UFCF x multiple)
        'pv_tv': 36,

        # Valuation
        'val_header': 38,
        'sum_pv_fcf': 39,
        'ev': 40,                   # = sum_pv + pv_tv
        'net_debt': 41,
        'equity_value': 42,         # = EV - net_debt
    },

    # =========================================================================
    # Comps (Sheet 12, ~45 rows)
    # =========================================================================
    'comps': {
        'header': 1,
        'table_header': 3,          # company, EV, Revenue, EBITDA, EBIT, NI, EV/Rev, EV/EBITDA, EV/EBIT, P/E
        'comp_1': 4,
        'comp_12': 15,              # 12 comps
        'blank_row': 16,
        'mean_row': 17,
        'median_row': 18,
        'q1_row': 19,
        'q3_row': 20,

        # Implied Valuation
        'implied_header': 22,
        'target_revenue': 23,
        'target_ebitda': 24,
        'target_ebit': 25,
        'target_ni': 26,
        'implied_ev_rev': 27,       # = target_revenue x median EV/Rev
        'implied_ev_ebitda': 28,
        'implied_ev_ebit': 29,
        'implied_pe': 30,
    },

    # =========================================================================
    # Dashboard (Sheet 13, ~30 rows)
    # =========================================================================
    'dashboard': {
        'header': 1,

        # KPI Summary
        'kpi_header': 3,
        'kpi_revenue': 4,
        'kpi_ebitda': 5,
        'kpi_ebitda_margin': 6,
        'kpi_net_income': 7,
        'kpi_net_margin': 8,
        'kpi_revenue_cagr': 9,

        # Returns Summary
        'returns_header': 11,
        'return_ps_irr': 12,
        'return_ps_moic': 13,
        'return_pe_irr': 14,
        'return_pe_moic': 15,

        # Model Checks
        'check_header': 17,
        'bs_check_status': 18,      # PASS/FAIL per year
        'cash_check_status': 19,

        # Key Metrics by Year
        'metrics_header': 21,
        'metrics_year_row': 22,
        'metrics_revenue': 23,
        'metrics_ebitda': 24,
        'metrics_ni': 25,
        'metrics_fcf': 26,
    },
}


# =============================================================================
# Helper Functions
# =============================================================================

def row(sheet_key, item_key):
    """
    Get the 1-based row number for a sheet item.

    Args:
        sheet_key: Key into ROWS dict (e.g., 'income_statement')
        item_key: Key within the sheet dict (e.g., 'revenue')

    Returns:
        int: 1-based row number

    Raises:
        KeyError: If sheet_key or item_key not found

    Example:
        >>> row('income_statement', 'revenue')
        5
    """
    return ROWS[sheet_key][item_key]


def cell_ref(sheet_key, item_key, col_idx):
    """
    Get a cell reference string like 'D5'.

    Args:
        sheet_key: Key into ROWS dict
        item_key: Key within the sheet dict
        col_idx: 1-based column index (e.g., 4 for column D)

    Returns:
        str: Cell reference (e.g., 'D5')

    Example:
        >>> cell_ref('balance_sheet', 'cash', 4)
        'D6'
    """
    r = ROWS[sheet_key][item_key]
    return f"{get_column_letter(col_idx)}{r}"


def sheet_cell_ref(sheet_key, item_key, col_idx):
    """
    Get a cross-sheet cell reference like "'Income Statement'!D30".

    Args:
        sheet_key: Key into ROWS/SHEETS dict
        item_key: Key within the sheet dict
        col_idx: 1-based column index

    Returns:
        str: Cross-sheet reference (e.g., "'Income Statement'!D30")

    Example:
        >>> sheet_cell_ref('income_statement', 'net_income', 8)
        "'Income Statement'!H30"
    """
    sheet_name = SHEETS[sheet_key]['name']
    r = ROWS[sheet_key][item_key]
    return f"'{sheet_name}'!{get_column_letter(col_idx)}{r}"


def data_range(sheet_key, item_key, start_col, end_col):
    """
    Get a range string like 'D5:L5' for a single row spanning multiple columns.

    Args:
        sheet_key: Key into ROWS dict
        item_key: Key within the sheet dict
        start_col: 1-based start column index
        end_col: 1-based end column index (inclusive)

    Returns:
        str: Range reference (e.g., 'D5:L5')

    Example:
        >>> data_range('income_statement', 'revenue', 4, 12)
        'D5:L5'
    """
    r = ROWS[sheet_key][item_key]
    return f"{get_column_letter(start_col)}{r}:{get_column_letter(end_col)}{r}"


def sheet_data_range(sheet_key, item_key, start_col, end_col):
    """
    Get a cross-sheet range string like "'Income Statement'!D5:L5".

    Args:
        sheet_key: Key into ROWS/SHEETS dict
        item_key: Key within the sheet dict
        start_col: 1-based start column index
        end_col: 1-based end column index (inclusive)

    Returns:
        str: Cross-sheet range reference

    Example:
        >>> sheet_data_range('income_statement', 'revenue', 4, 12)
        "'Income Statement'!D5:L5"
    """
    sheet_name = SHEETS[sheet_key]['name']
    r = ROWS[sheet_key][item_key]
    return f"'{sheet_name}'!{get_column_letter(start_col)}{r}:{get_column_letter(end_col)}{r}"


# =============================================================================
# Validation — ensure no duplicate rows within a sheet
# =============================================================================

def _validate_row_map():
    """
    Validate the ROWS map on import. Raises ValueError if any sheet has
    duplicate row numbers (a sign of copy-paste errors).
    """
    for sheet_key, items in ROWS.items():
        seen = {}
        for item_key, row_num in items.items():
            if not isinstance(row_num, int) or row_num < 1:
                raise ValueError(
                    f"ROWS['{sheet_key}']['{item_key}'] = {row_num!r} "
                    f"is not a valid positive integer row number"
                )
            if row_num in seen:
                raise ValueError(
                    f"ROWS['{sheet_key}'] has duplicate row {row_num}: "
                    f"'{seen[row_num]}' and '{item_key}'"
                )
            seen[row_num] = item_key


# Run validation on import — fail fast if the map is corrupted
_validate_row_map()
