"""
Cathay Capital PE Financial Model — Template Builder.

Generates a 13-sheet Cathay-branded Excel financial model template.
Uses row_map for all row positions, formula_engine for all calculations,
and format_engine for all visual formatting.

Usage:
    from template_builder import build_template
    path = build_template()  # -> '/path/to/cathay_pe_model.xlsx'
"""

import sys
import os

_LIB_DIR = os.path.dirname(os.path.abspath(__file__))
if _LIB_DIR not in sys.path:
    sys.path.insert(0, _LIB_DIR)

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from constants import (
    SHEET_NAMES, NUM_SHEETS,
    COL_LABEL, COL_UNIT, COL_NOTE,
    COL_HIST_START, COL_HIST_END,
    COL_FORECAST_START, COL_FORECAST_END, COL_LAST,
    ALL_YEARS, HIST_YEARS, FORECAST_YEARS,
    NUM_FMT_NUMBER, NUM_FMT_PCT, NUM_FMT_MULTIPLE, NUM_FMT_YEAR,
    INPUT_FONT, FORMULA_FONT, LINK_FONT,
    HEADER_FONT, HEADER_FILL, ALIGN_HEADER,
    CATHAY_RED, CATHAY_GOLD,
)
from row_map import ROWS, SHEETS
from formula_engine import get_all_formulas
from format_engine import (
    apply_column_widths, apply_header_row, apply_subheader_row,
    apply_year_row, apply_label, apply_total_row_format,
    apply_subtotal_row_format, apply_check_row_format,
    apply_number_format, apply_hist_forecast_divider,
    apply_alternating_rows, apply_input_font, apply_formula_font,
    apply_link_font, format_sheet,
)


# =============================================================================
# Label Definitions (per sheet)
# =============================================================================

# Each entry: (row_map_key, label_text, unit, options_dict)
# options_dict can contain: bold, indent, is_header, is_total, is_check, is_pct, is_multiple

COVER_LABELS = [
    ('company_name', 'Company Name', ''),
    ('industry', 'Industry / Sector', ''),
    ('date', 'Date', ''),
    ('analyst', 'Analyst', ''),
    ('fx_rate', 'FX Rate (USD/RMB)', ''),
    ('base_currency', 'Base Currency', ''),
]

ASSUMPTIONS_LABELS = [
    # Scenario
    ('scenario_label', 'Scenario', ''),
    ('scenario_toggle', 'Active Scenario', ''),
    ('revenue_multiplier', 'Revenue Multiplier', 'x'),
    ('margin_adjustment', 'Margin Adjustment', 'bps'),
    # Segment A
    ('seg_a_header', 'SEGMENT A REVENUE DRIVERS', ''),
    ('seg_a_volume', 'Volume (units)', 'units'),
    ('seg_a_price', 'ASP', 'RMB'),
    ('seg_a_utilization', 'Utilization Rate', '%'),
    ('seg_a_revenue', 'Segment A Revenue', 'RMB mn'),
    ('seg_a_growth', '  YoY Growth', '%'),
    ('seg_a_pct', '  % of Total Revenue', '%'),
    # Segment B
    ('seg_b_header', 'SEGMENT B REVENUE DRIVERS', ''),
    ('seg_b_volume', 'Volume (units)', 'units'),
    ('seg_b_price', 'ASP', 'RMB'),
    ('seg_b_utilization', 'Utilization Rate', '%'),
    ('seg_b_revenue', 'Segment B Revenue', 'RMB mn'),
    ('seg_b_growth', '  YoY Growth', '%'),
    ('seg_b_pct', '  % of Total Revenue', '%'),
    # Segment C
    ('seg_c_header', 'SEGMENT C REVENUE DRIVERS', ''),
    ('seg_c_volume', 'Volume (units)', 'units'),
    ('seg_c_price', 'ASP', 'RMB'),
    ('seg_c_utilization', 'Utilization Rate', '%'),
    ('seg_c_revenue', 'Segment C Revenue', 'RMB mn'),
    ('seg_c_growth', '  YoY Growth', '%'),
    ('seg_c_pct', '  % of Total Revenue', '%'),
    # Total
    ('total_revenue', 'Total Revenue', 'RMB mn'),
    ('total_growth', '  Total Revenue Growth', '%'),
    # Cost
    ('cost_header', 'COST ASSUMPTIONS', ''),
    ('cogs_pct_a', 'COGS % — Segment A', '%'),
    ('cogs_pct_b', 'COGS % — Segment B', '%'),
    ('cogs_pct_c', 'COGS % — Segment C', '%'),
    ('blended_cogs_pct', 'Blended COGS %', '%'),
    ('gross_margin', '  Gross Margin', '%'),
    # OpEx
    ('opex_header', 'OPEX ASSUMPTIONS', ''),
    ('personnel_pct', 'Personnel (% of Rev)', '%'),
    ('rent_pct', 'Rent (% of Rev)', '%'),
    ('marketing_pct', 'Marketing (% of Rev)', '%'),
    ('rd_pct', 'R&D (% of Rev)', '%'),
    ('other_opex_pct', 'Other OpEx (% of Rev)', '%'),
    ('total_sga_pct', 'Total SG&A %', '%'),
    # CapEx
    ('capex_header', 'CAPEX & D&A ASSUMPTIONS', ''),
    ('capex_pct', 'CapEx (% of Rev)', '%'),
    ('da_rate', 'D&A Rate (% of PP&E)', '%'),
    # WC
    ('wc_header', 'WORKING CAPITAL ASSUMPTIONS', ''),
    ('ar_days', 'AR Days', 'days'),
    ('ap_days', 'AP Days', 'days'),
    ('inventory_days', 'Inventory Days', 'days'),
    # Tax
    ('tax_header', 'TAX & OTHER', ''),
    ('tax_rate', 'Effective Tax Rate', '%'),
    ('dividend_payout', 'Dividend Payout Ratio', '%'),
    ('interest_rate', 'Interest Rate (avg)', '%'),
]

REVENUE_LABELS = [
    ('bu_header', 'BOTTOM-UP REVENUE BUILD', ''),
    ('bu_seg_a', 'Segment A', 'RMB mn'),
    ('bu_seg_b', 'Segment B', 'RMB mn'),
    ('bu_seg_c', 'Segment C', 'RMB mn'),
    ('bu_total', 'Total Revenue (Bottom-Up)', 'RMB mn'),
    ('td_header', 'TOP-DOWN CROSS-CHECK', ''),
    ('tam', 'Total Addressable Market', 'RMB mn'),
    ('penetration', 'Market Penetration', '%'),
    ('market_share', 'Market Share', '%'),
    ('td_total', 'Total Revenue (Top-Down)', 'RMB mn'),
    ('recon_header', 'RECONCILIATION', ''),
    ('recon_diff', 'BU vs TD Variance', '%'),
    ('recon_flag', 'Flag (>10% = CHECK)', ''),
    ('mix_header', 'REVENUE MIX', ''),
    ('mix_seg_a_pct', 'Segment A %', '%'),
    ('mix_seg_b_pct', 'Segment B %', '%'),
    ('mix_seg_c_pct', 'Segment C %', '%'),
    ('growth_header', 'YOY GROWTH', ''),
    ('growth_total', 'Total Revenue Growth', '%'),
    ('growth_seg_a', 'Segment A Growth', '%'),
    ('growth_seg_b', 'Segment B Growth', '%'),
    ('growth_seg_c', 'Segment C Growth', '%'),
]

COGS_OPEX_LABELS = [
    ('cogs_header', 'COST OF GOODS SOLD', ''),
    ('cogs_seg_a', 'COGS — Segment A', 'RMB mn'),
    ('cogs_seg_b', 'COGS — Segment B', 'RMB mn'),
    ('cogs_seg_c', 'COGS — Segment C', 'RMB mn'),
    ('total_cogs', 'Total COGS', 'RMB mn'),
    ('cogs_pct_rev', '  COGS % of Revenue', '%'),
    ('gross_profit', 'Gross Profit', 'RMB mn'),
    ('gross_margin', '  Gross Margin', '%'),
    ('sga_header', 'SG&A EXPENSES', ''),
    ('personnel', 'Personnel', 'RMB mn'),
    ('rent', 'Rent & Facilities', 'RMB mn'),
    ('marketing', 'Marketing & Sales', 'RMB mn'),
    ('rd', 'Research & Development', 'RMB mn'),
    ('other_opex', 'Other Operating Expenses', 'RMB mn'),
    ('total_sga', 'Total SG&A', 'RMB mn'),
    ('sga_pct_rev', '  SG&A % of Revenue', '%'),
    ('da_header', 'DEPRECIATION & AMORTIZATION', ''),
    ('depreciation', 'Depreciation', 'RMB mn'),
    ('amortization', 'Amortization', 'RMB mn'),
    ('total_da', 'Total D&A', 'RMB mn'),
    ('total_opex', 'Total Operating Expenses', 'RMB mn'),
    ('ebitda_bridge_header', 'EBITDA BRIDGE', ''),
    ('ebitda', 'EBITDA', 'RMB mn'),
    ('ebitda_margin', '  EBITDA Margin', '%'),
    ('ebit', 'EBIT', 'RMB mn'),
    ('ebit_margin', '  EBIT Margin', '%'),
]

INCOME_STATEMENT_LABELS = [
    ('revenue', 'Revenue', 'RMB mn'),
    ('cogs', 'Cost of Goods Sold', 'RMB mn'),
    ('gross_profit', 'Gross Profit', 'RMB mn'),
    ('gross_margin', '  Gross Margin', '%'),
    ('sga', 'SG&A Expenses', 'RMB mn'),
    ('other_income', 'Other Income / (Expense)', 'RMB mn'),
    ('ebitda', 'EBITDA', 'RMB mn'),
    ('ebitda_margin', '  EBITDA Margin', '%'),
    ('da', 'Depreciation & Amortization', 'RMB mn'),
    ('ebit', 'EBIT', 'RMB mn'),
    ('ebit_margin', '  EBIT Margin', '%'),
    ('interest_expense', 'Interest Expense', 'RMB mn'),
    ('other_fin', 'Other Financial Items', 'RMB mn'),
    ('ebt', 'Earnings Before Tax', 'RMB mn'),
    ('tax', 'Income Tax', 'RMB mn'),
    ('effective_tax_rate', '  Effective Tax Rate', '%'),
    ('net_income', 'Net Income', 'RMB mn'),
    ('net_margin', '  Net Margin', '%'),
    ('shares_outstanding', 'Shares Outstanding', 'mn'),
    ('eps', 'Earnings Per Share', 'RMB'),
]

BALANCE_SHEET_LABELS = [
    ('ca_header', 'CURRENT ASSETS', ''),
    ('cash', 'Cash & Equivalents', 'RMB mn'),
    ('accounts_receivable', 'Accounts Receivable', 'RMB mn'),
    ('inventory', 'Inventory', 'RMB mn'),
    ('other_current', 'Other Current Assets', 'RMB mn'),
    ('total_current_assets', 'Total Current Assets', 'RMB mn'),
    ('nca_header', 'NON-CURRENT ASSETS', ''),
    ('ppe_net', 'PP&E, Net', 'RMB mn'),
    ('intangibles', 'Intangibles & Goodwill', 'RMB mn'),
    ('other_noncurrent', 'Other Non-Current Assets', 'RMB mn'),
    ('total_noncurrent_assets', 'Total Non-Current Assets', 'RMB mn'),
    ('total_assets', 'TOTAL ASSETS', 'RMB mn'),
    ('cl_header', 'CURRENT LIABILITIES', ''),
    ('accounts_payable', 'Accounts Payable', 'RMB mn'),
    ('st_debt', 'Short-term Debt', 'RMB mn'),
    ('other_current_liab', 'Other Current Liabilities', 'RMB mn'),
    ('total_current_liab', 'Total Current Liabilities', 'RMB mn'),
    ('ncl_header', 'NON-CURRENT LIABILITIES', ''),
    ('lt_debt', 'Long-term Debt', 'RMB mn'),
    ('other_noncurrent_liab', 'Other Non-Current Liabilities', 'RMB mn'),
    ('total_noncurrent_liab', 'Total Non-Current Liabilities', 'RMB mn'),
    ('total_liabilities', 'Total Liabilities', 'RMB mn'),
    ('eq_header', 'EQUITY', ''),
    ('paid_in_capital', 'Paid-in Capital', 'RMB mn'),
    ('retained_earnings', 'Retained Earnings', 'RMB mn'),
    ('total_equity', 'Total Equity', 'RMB mn'),
    ('total_le', 'TOTAL LIABILITIES + EQUITY', 'RMB mn'),
    ('bs_check', 'Balance Check (must = 0)', 'RMB mn'),
]

CASH_FLOW_LABELS = [
    ('op_header', 'OPERATING ACTIVITIES', ''),
    ('net_income', 'Net Income', 'RMB mn'),
    ('da', 'Depreciation & Amortization', 'RMB mn'),
    ('delta_ar', 'Change in Accounts Receivable', 'RMB mn'),
    ('delta_inventory', 'Change in Inventory', 'RMB mn'),
    ('delta_ap', 'Change in Accounts Payable', 'RMB mn'),
    ('delta_other_wc', 'Change in Other Working Capital', 'RMB mn'),
    ('total_delta_wc', 'Total Change in Working Capital', 'RMB mn'),
    ('operating_cf', 'Cash from Operations', 'RMB mn'),
    ('inv_header', 'INVESTING ACTIVITIES', ''),
    ('capex', 'Capital Expenditures', 'RMB mn'),
    ('other_investing', 'Other Investing Activities', 'RMB mn'),
    ('investing_cf', 'Cash from Investing', 'RMB mn'),
    ('fin_header', 'FINANCING ACTIVITIES', ''),
    ('debt_drawdown', 'Debt Drawdown', 'RMB mn'),
    ('debt_repayment', 'Debt Repayment', 'RMB mn'),
    ('dividends', 'Dividends Paid', 'RMB mn'),
    ('equity_issuance', 'Equity Issuance / (Buyback)', 'RMB mn'),
    ('financing_cf', 'Cash from Financing', 'RMB mn'),
    ('net_cf', 'Net Change in Cash', 'RMB mn'),
    ('beginning_cash', 'Beginning Cash', 'RMB mn'),
    ('ending_cash', 'Ending Cash', 'RMB mn'),
    ('cash_tieout', 'Cash Tie-out (must = 0)', 'RMB mn'),
]

WORKING_CAPITAL_LABELS = [
    ('days_header', 'WORKING CAPITAL DAYS', ''),
    ('ar_days', 'Accounts Receivable Days', 'days'),
    ('inventory_days', 'Inventory Days', 'days'),
    ('ap_days', 'Accounts Payable Days', 'days'),
    ('bal_header', 'WORKING CAPITAL BALANCES', ''),
    ('ar_balance', 'Accounts Receivable', 'RMB mn'),
    ('inventory_balance', 'Inventory', 'RMB mn'),
    ('ap_balance', 'Accounts Payable', 'RMB mn'),
    ('net_wc', 'Net Working Capital', 'RMB mn'),
    ('change_header', 'CHANGES IN WORKING CAPITAL', ''),
    ('delta_ar', 'Change in AR', 'RMB mn'),
    ('delta_inventory', 'Change in Inventory', 'RMB mn'),
    ('delta_ap', 'Change in AP', 'RMB mn'),
    ('delta_net_wc', 'Change in Net Working Capital', 'RMB mn'),
]

DEBT_CAPEX_LABELS = [
    ('d1_header', 'DEBT TRANCHE 1', ''),
    ('d1_beginning', 'Beginning Balance', 'RMB mn'),
    ('d1_drawdown', 'Drawdown', 'RMB mn'),
    ('d1_repayment', 'Repayment', 'RMB mn'),
    ('d1_ending', 'Ending Balance', 'RMB mn'),
    ('d1_rate', 'Interest Rate', '%'),
    ('d1_interest', 'Interest Expense', 'RMB mn'),
    ('d2_header', 'DEBT TRANCHE 2', ''),
    ('d2_beginning', 'Beginning Balance', 'RMB mn'),
    ('d2_drawdown', 'Drawdown', 'RMB mn'),
    ('d2_repayment', 'Repayment', 'RMB mn'),
    ('d2_ending', 'Ending Balance', 'RMB mn'),
    ('d2_rate', 'Interest Rate', '%'),
    ('d2_interest', 'Interest Expense', 'RMB mn'),
    ('debt_summary_header', 'DEBT SUMMARY', ''),
    ('total_debt', 'Total Debt Outstanding', 'RMB mn'),
    ('total_interest', 'Total Interest Expense', 'RMB mn'),
    ('total_drawdown', 'Total Drawdown', 'RMB mn'),
    ('total_repayment', 'Total Repayment', 'RMB mn'),
    ('capex_header', 'CAPITAL EXPENDITURES', ''),
    ('capex_maintenance', 'Maintenance CapEx', 'RMB mn'),
    ('capex_growth', 'Growth CapEx', 'RMB mn'),
    ('total_capex', 'Total CapEx', 'RMB mn'),
    ('da_header', 'PP&E ROLL-FORWARD', ''),
    ('beginning_ppe', 'Beginning PP&E', 'RMB mn'),
    ('plus_capex', '(+) CapEx', 'RMB mn'),
    ('less_da', '(-) Depreciation', 'RMB mn'),
    ('ending_ppe', 'Ending PP&E', 'RMB mn'),
]

RETURNS_LABELS = [
    ('entry_header', 'ENTRY ASSUMPTIONS', ''),
    ('entry_revenue', 'Entry Revenue', 'RMB mn'),
    ('entry_net_income', 'Entry Net Income', 'RMB mn'),
    ('entry_ps_multiple', 'Entry P/S Multiple', 'x'),
    ('entry_pe_multiple', 'Entry P/E Multiple', 'x'),
    ('entry_ev', 'Entry Enterprise Value', 'RMB mn'),
    ('entry_net_debt', 'Entry Net Debt', 'RMB mn'),
    ('entry_equity', 'Entry Equity Value', 'RMB mn'),
    ('ps_exit_header', 'EXIT — P/S METHOD', ''),
    ('ps_exit_revenue', 'Exit Year Revenue', 'RMB mn'),
    ('ps_exit_multiple', 'Exit P/S Multiple', 'x'),
    ('ps_exit_ev', 'Exit Enterprise Value', 'RMB mn'),
    ('ps_exit_net_debt', 'Exit Net Debt', 'RMB mn'),
    ('ps_exit_equity', 'Exit Equity Value', 'RMB mn'),
    ('ps_moic', 'MOIC (P/S)', 'x'),
    ('ps_irr', 'IRR (P/S)', '%'),
    ('pe_exit_header', 'EXIT — P/E METHOD', ''),
    ('pe_exit_ni', 'Exit Year Net Income', 'RMB mn'),
    ('pe_exit_multiple', 'Exit P/E Multiple', 'x'),
    ('pe_exit_equity', 'Exit Equity Value', 'RMB mn'),
    ('pe_moic', 'MOIC (P/E)', 'x'),
    ('pe_irr', 'IRR (P/E)', '%'),
    ('sens1_header', 'SENSITIVITY: IRR vs ENTRY x EXIT P/E', ''),
    ('sens2_header', 'SENSITIVITY: MOIC vs REV CAGR x EXIT YEAR', ''),
]

DCF_LABELS = [
    ('wacc_header', 'WACC CALCULATION', ''),
    ('risk_free', 'Risk-Free Rate', '%'),
    ('beta', 'Beta', 'x'),
    ('mrp', 'Market Risk Premium', '%'),
    ('cost_of_equity', 'Cost of Equity', '%'),
    ('cost_of_debt', 'Pre-tax Cost of Debt', '%'),
    ('tax_rate', 'Tax Rate', '%'),
    ('equity_weight', 'Equity Weight (E/V)', '%'),
    ('debt_weight', 'Debt Weight (D/V)', '%'),
    ('wacc', 'WACC', '%'),
    ('ufcf_header', 'UNLEVERED FREE CASH FLOW', ''),
    ('ebit', 'EBIT', 'RMB mn'),
    ('tax_on_ebit', 'Tax on EBIT', 'RMB mn'),
    ('nopat', 'NOPAT', 'RMB mn'),
    ('plus_da', '(+) D&A', 'RMB mn'),
    ('less_capex', '(-) CapEx', 'RMB mn'),
    ('less_delta_wc', '(-) Change in WC', 'RMB mn'),
    ('ufcf', 'Unlevered FCF', 'RMB mn'),
    ('discount_header', 'DISCOUNT FACTORS', ''),
    ('discount_period', 'Discount Period (mid-year)', ''),
    ('discount_factor', 'Discount Factor', ''),
    ('pv_fcf', 'PV of FCF', 'RMB mn'),
    ('tv_header', 'TERMINAL VALUE', ''),
    ('tv_method_toggle', 'TV Method', ''),
    ('terminal_growth', 'Terminal Growth Rate', '%'),
    ('exit_multiple', 'Exit EV/EBITDA Multiple', 'x'),
    ('terminal_value', 'Terminal Value', 'RMB mn'),
    ('pv_tv', 'PV of Terminal Value', 'RMB mn'),
    ('val_header', 'ENTERPRISE & EQUITY VALUE', ''),
    ('sum_pv_fcf', 'Sum of PV(FCF)', 'RMB mn'),
    ('ev', 'Enterprise Value', 'RMB mn'),
    ('net_debt', 'Net Debt', 'RMB mn'),
    ('equity_value', 'Equity Value', 'RMB mn'),
]

COMPS_LABELS = [
    ('table_header', 'COMPARABLE COMPANIES', ''),
    ('comp_1', 'Comp 1', 'RMB mn'),
    # comps 2-12 handled in loop
    ('mean_row', 'Mean', ''),
    ('median_row', 'Median', ''),
    ('q1_row', '25th Percentile', ''),
    ('q3_row', '75th Percentile', ''),
    ('implied_header', 'IMPLIED VALUATION', ''),
    ('target_revenue', 'Target Revenue (NTM)', 'RMB mn'),
    ('target_ebitda', 'Target EBITDA (NTM)', 'RMB mn'),
    ('target_ebit', 'Target EBIT (NTM)', 'RMB mn'),
    ('target_ni', 'Target Net Income (NTM)', 'RMB mn'),
    ('implied_ev_rev', 'Implied EV (EV/Revenue)', 'RMB mn'),
    ('implied_ev_ebitda', 'Implied EV (EV/EBITDA)', 'RMB mn'),
    ('implied_ev_ebit', 'Implied EV (EV/EBIT)', 'RMB mn'),
    ('implied_pe', 'Implied Equity (P/E)', 'RMB mn'),
]

DASHBOARD_LABELS = [
    ('kpi_header', 'KEY PERFORMANCE INDICATORS', ''),
    ('kpi_revenue', 'Revenue', 'RMB mn'),
    ('kpi_ebitda', 'EBITDA', 'RMB mn'),
    ('kpi_ebitda_margin', 'EBITDA Margin', '%'),
    ('kpi_net_income', 'Net Income', 'RMB mn'),
    ('kpi_net_margin', 'Net Margin', '%'),
    ('kpi_revenue_cagr', 'Revenue CAGR (5yr)', '%'),
    ('returns_header', 'RETURNS SUMMARY', ''),
    ('return_ps_irr', 'IRR (P/S Method)', '%'),
    ('return_ps_moic', 'MOIC (P/S Method)', 'x'),
    ('return_pe_irr', 'IRR (P/E Method)', '%'),
    ('return_pe_moic', 'MOIC (P/E Method)', 'x'),
    ('check_header', 'MODEL INTEGRITY CHECKS', ''),
    ('bs_check_status', 'Balance Sheet Check', ''),
    ('cash_check_status', 'Cash Tie-out Check', ''),
    ('metrics_header', 'KEY METRICS BY YEAR', ''),
    ('metrics_year_row', '', ''),
    ('metrics_revenue', 'Revenue', 'RMB mn'),
    ('metrics_ebitda', 'EBITDA', 'RMB mn'),
    ('metrics_ni', 'Net Income', 'RMB mn'),
    ('metrics_fcf', 'Free Cash Flow', 'RMB mn'),
]

# Map sheet keys to their label definitions
LABEL_DEFS = {
    'cover': COVER_LABELS,
    'assumptions': ASSUMPTIONS_LABELS,
    'revenue': REVENUE_LABELS,
    'cogs_opex': COGS_OPEX_LABELS,
    'income_statement': INCOME_STATEMENT_LABELS,
    'balance_sheet': BALANCE_SHEET_LABELS,
    'cash_flow': CASH_FLOW_LABELS,
    'working_capital': WORKING_CAPITAL_LABELS,
    'debt_capex': DEBT_CAPEX_LABELS,
    'returns': RETURNS_LABELS,
    'dcf': DCF_LABELS,
    'comps': COMPS_LABELS,
    'dashboard': DASHBOARD_LABELS,
}

# Subheader keys (rows that get subheader formatting, not labels)
SUBHEADER_KEYS = {
    'assumptions': ['seg_a_header', 'seg_b_header', 'seg_c_header',
                    'cost_header', 'opex_header', 'capex_header', 'wc_header', 'tax_header'],
    'revenue': ['bu_header', 'td_header', 'recon_header', 'mix_header', 'growth_header'],
    'cogs_opex': ['cogs_header', 'sga_header', 'da_header', 'ebitda_bridge_header'],
    'income_statement': [],
    'balance_sheet': ['ca_header', 'nca_header', 'cl_header', 'ncl_header', 'eq_header'],
    'cash_flow': ['op_header', 'inv_header', 'fin_header'],
    'working_capital': ['days_header', 'bal_header', 'change_header'],
    'debt_capex': ['d1_header', 'd2_header', 'debt_summary_header', 'capex_header', 'da_header'],
    'returns': ['entry_header', 'ps_exit_header', 'pe_exit_header', 'sens1_header', 'sens2_header'],
    'dcf': ['wacc_header', 'ufcf_header', 'discount_header', 'tv_header', 'val_header'],
    'comps': ['implied_header'],
    'dashboard': ['kpi_header', 'returns_header', 'check_header', 'metrics_header'],
}

# Total rows (bold + double border)
TOTAL_KEYS = {
    'assumptions': ['total_revenue'],
    'revenue': ['bu_total', 'td_total'],
    'cogs_opex': ['total_cogs', 'gross_profit', 'total_sga', 'total_da', 'total_opex', 'ebitda', 'ebit'],
    'income_statement': ['gross_profit', 'ebitda', 'ebit', 'ebt', 'net_income'],
    'balance_sheet': ['total_current_assets', 'total_noncurrent_assets', 'total_assets',
                      'total_current_liab', 'total_noncurrent_liab', 'total_liabilities',
                      'total_equity', 'total_le'],
    'cash_flow': ['operating_cf', 'investing_cf', 'financing_cf', 'net_cf', 'ending_cash'],
    'working_capital': ['net_wc', 'delta_net_wc'],
    'debt_capex': ['total_debt', 'total_interest', 'total_capex', 'ending_ppe'],
    'returns': ['ps_exit_equity', 'pe_exit_equity'],
    'dcf': ['ufcf', 'ev', 'equity_value'],
    'comps': ['mean_row', 'median_row'],
    'dashboard': [],
}

# Check rows (pass/fail)
CHECK_KEYS = {
    'balance_sheet': ['bs_check'],
    'cash_flow': ['cash_tieout'],
    'dashboard': ['bs_check_status', 'cash_check_status'],
}

# Rows that get link font (green) in forecast columns
LINK_KEYS = {
    'income_statement': ['revenue', 'cogs', 'sga', 'da', 'interest_expense', 'tax'],
    'balance_sheet': ['cash', 'accounts_receivable', 'inventory', 'ppe_net',
                      'accounts_payable', 'lt_debt', 'retained_earnings'],
    'cash_flow': ['net_income', 'da', 'delta_ar', 'delta_inventory', 'delta_ap',
                  'capex', 'debt_drawdown', 'debt_repayment', 'dividends'],
    'working_capital': ['ar_balance', 'inventory_balance', 'ap_balance'],
    'cogs_opex': ['cogs_seg_a', 'cogs_seg_b', 'cogs_seg_c',
                  'personnel', 'rent', 'marketing', 'rd', 'other_opex', 'depreciation'],
    'dcf': ['ebit', 'plus_da', 'less_capex', 'less_delta_wc'],
    'comps': ['target_revenue', 'target_ebitda', 'target_ebit', 'target_ni'],
    'dashboard': ['kpi_revenue', 'kpi_ebitda', 'kpi_net_income',
                  'metrics_revenue', 'metrics_ebitda', 'metrics_ni', 'metrics_fcf'],
}


# =============================================================================
# Builder
# =============================================================================

def _write_labels(ws, sheet_key):
    """Write row labels (col A) and units (col B) for a sheet."""
    labels = LABEL_DEFS.get(sheet_key, [])
    r = ROWS.get(sheet_key, {})
    subheaders = SUBHEADER_KEYS.get(sheet_key, [])

    for item in labels:
        key, text, unit = item[0], item[1], item[2]
        if key not in r:
            continue
        row_num = r[key]

        if key in subheaders:
            apply_subheader_row(ws, row_num, text)
        else:
            # Determine if bold (total rows)
            is_bold = key in TOTAL_KEYS.get(sheet_key, [])
            # Determine indent (labels starting with spaces)
            indent = 0
            if text.startswith('  '):
                indent = 1
                text = text.strip()

            apply_label(ws, row_num, text, unit=unit, bold=is_bold, indent=indent)

    # Handle comps rows 2-12 (not in COMPS_LABELS to avoid clutter)
    if sheet_key == 'comps':
        comp_start = r.get('comp_1', 4)
        comp_end = r.get('comp_12', 15)
        for i, row_num in enumerate(range(comp_start + 1, comp_end + 1)):
            apply_label(ws, row_num, f'Comp {i + 2}', unit='RMB mn')


def _write_year_row(ws, sheet_key):
    """Write the year header row (2021-2029E)."""
    r = ROWS.get(sheet_key, {})
    if 'year_row' in r:
        apply_year_row(ws, r['year_row'])


def _write_formulas(ws, sheet_key, all_formulas):
    """Write all formulas for a sheet into forecast columns."""
    sheet_formulas = all_formulas.get(sheet_key, {})
    for (row_num, col), formula in sheet_formulas.items():
        cell = ws.cell(row=row_num, column=col)
        if isinstance(formula, str):
            cell.value = formula
        else:
            # Numeric constant (e.g., discount_period mid-year value)
            cell.value = formula


def _apply_fonts(ws, sheet_key):
    """Apply input (blue), formula (black), and link (green) fonts."""
    r = ROWS.get(sheet_key, {})
    subheaders = SUBHEADER_KEYS.get(sheet_key, [])
    link_keys = LINK_KEYS.get(sheet_key, [])

    for key, row_num in r.items():
        if key in ('header', 'year_row') or key in subheaders:
            continue

        # Historical columns: INPUT_FONT (blue)
        apply_input_font(ws, row_num, COL_HIST_START, COL_HIST_END)

        # Forecast columns: determine font
        if key in link_keys:
            apply_link_font(ws, row_num, COL_FORECAST_START, COL_FORECAST_END)
        else:
            apply_formula_font(ws, row_num, COL_FORECAST_START, COL_FORECAST_END)


def _apply_formatting(ws, sheet_key):
    """Apply all formatting: totals, checks, number formats, divider, alternating."""
    r = ROWS.get(sheet_key, {})

    # Total rows
    for key in TOTAL_KEYS.get(sheet_key, []):
        if key in r:
            apply_total_row_format(ws, r[key])

    # Check rows
    for key in CHECK_KEYS.get(sheet_key, []):
        if key in r:
            apply_check_row_format(ws, r[key])

    # Number formats from label definitions
    labels = LABEL_DEFS.get(sheet_key, [])
    for item in labels:
        key, text, unit = item[0], item[1], item[2]
        if key not in r:
            continue
        row_num = r[key]
        if unit == '%':
            apply_number_format(ws, row_num, NUM_FMT_PCT)
        elif unit == 'x':
            apply_number_format(ws, row_num, NUM_FMT_MULTIPLE)
        elif unit in ('RMB mn', 'RMB', 'mn', 'units'):
            apply_number_format(ws, row_num, NUM_FMT_NUMBER)
        elif unit == 'days':
            apply_number_format(ws, row_num, NUM_FMT_NUMBER)

    # Hist/forecast divider
    if 'year_row' in r:
        max_row = max(r.values())
        apply_hist_forecast_divider(ws, r['year_row'], max_row)

    # Alternating rows
    if 'year_row' in r:
        start = r['year_row'] + 1
        max_row = max(r.values())
        apply_alternating_rows(ws, start, max_row)


def _set_print_area(ws, sheet_key):
    """Set print area to cover all data columns."""
    r = ROWS.get(sheet_key, {})
    if not r:
        return
    max_row = max(r.values())
    ws.print_area = f"A1:{get_column_letter(COL_LAST)}{max_row}"


def _add_scenario_dropdown(ws):
    """Add scenario dropdown DataValidation on Key Assumptions sheet."""
    r = ROWS['assumptions']
    toggle_row = r['scenario_toggle']
    # Apply to forecast start column (H)
    dv = DataValidation(
        type="list",
        formula1='"Base,Upside,Downside"',
        allow_blank=True,
    )
    dv.error = "Please select Base, Upside, or Downside"
    dv.errorTitle = "Invalid Scenario"
    cell_ref = f"{get_column_letter(COL_FORECAST_START)}{toggle_row}"
    dv.add(cell_ref)
    ws.add_data_validation(dv)
    # Set default value
    ws.cell(row=toggle_row, column=COL_FORECAST_START).value = "Base"


def build_template(output_path=None):
    """Generate a fresh 13-sheet PE financial model template.

    Args:
        output_path: Optional path for the output .xlsx file.
                     Defaults to 'cathay_pe_model.xlsx' in current directory.

    Returns:
        str: Path to generated .xlsx file.
    """
    if output_path is None:
        output_path = os.path.join(os.getcwd(), 'cathay_pe_model.xlsx')

    wb = Workbook()

    # Get all formulas upfront
    all_formulas = get_all_formulas()

    # Create all 13 sheets
    for sheet_name in SHEET_NAMES:
        wb.create_sheet(title=sheet_name)

    # Delete the default "Sheet" created by Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Build each sheet
    for sheet_key, sheet_meta in SHEETS.items():
        sheet_name = sheet_meta['name']
        ws = wb[sheet_name]

        # 1. Column widths
        apply_column_widths(ws)

        # 2. Header row (row 1 for all sheets with a 'header' key)
        r = ROWS.get(sheet_key, {})
        if 'header' in r:
            apply_header_row(ws, r['header'], sheet_name)

        # 3. Year row
        _write_year_row(ws, sheet_key)

        # 4. Labels in column A, units in column B
        _write_labels(ws, sheet_key)

        # 5. Formulas in forecast columns
        _write_formulas(ws, sheet_key, all_formulas)

        # 6. Fonts (input/formula/link)
        _apply_fonts(ws, sheet_key)

        # 7. Formatting (totals, checks, number formats, divider, alternating)
        _apply_formatting(ws, sheet_key)

        # 8. Print area
        _set_print_area(ws, sheet_key)

    # Scenario dropdown on Key Assumptions
    assumptions_ws = wb['Key Assumptions']
    _add_scenario_dropdown(assumptions_ws)

    # Save workbook
    wb.save(output_path)
    return output_path


# =============================================================================
# CLI Entry Point
# =============================================================================

if __name__ == '__main__':
    path = build_template()
    from openpyxl import load_workbook
    wb = load_workbook(path)
    print(f"Generated: {path}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row}r x {ws.max_column}c")
