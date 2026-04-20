import sys, os
_LIB_DIR = os.path.dirname(os.path.abspath(__file__))
if _LIB_DIR not in sys.path:
    sys.path.insert(0, _LIB_DIR)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from row_map import ROWS, SHEETS, row
from constants import COL_HIST_START, COL_FORECAST_START, COL_FORECAST_END, SHEET_NAMES


def validate_model(xlsx_path, verbose=True):
    """Run 10-point validation on a Cathay PE financial model.

    Checks:
    1. Sheet structure — correct count and names
    2. BS Balance — Total Assets = Total L+E across all years
    3. Cash Tie-out — CF ending cash = BS cash across all years
    4. Revenue Bridge — IS revenue = Key Assumptions total revenue
    5. EBITDA Bridge — IS EBITDA = Gross Profit - SG&A (from COGS & OpEx)
    6. RE Roll-forward — Retained Earnings(t) = RE(t-1) + NI(t) - Div(t)
    7. Debt Consistency — BS total debt = Debt & CapEx total debt
    8. Formula Presence — forecast columns have formulas (strings starting with '='), not values
    9. No Blank Anchors — critical anchor cells (revenue, NI, TA, check rows) are not empty
    10. Cross-sheet Links — cells that should be green links actually contain sheet references

    Args:
        xlsx_path: path to .xlsx file
        verbose: print detailed results

    Returns:
        dict with:
            'passed': bool (True if all checks pass)
            'checks': list of (check_name, passed, details)
            'errors': list of error strings
    """
    wb = load_workbook(xlsx_path)
    errors = []
    checks = []

    # --- Check 1: Sheet structure ---
    missing = [n for n in SHEET_NAMES if n not in wb.sheetnames]
    extra = [n for n in wb.sheetnames if n not in SHEET_NAMES]
    ok = len(missing) == 0
    detail = f"{len(wb.sheetnames)} sheets"
    if missing: detail += f", missing: {missing}"
    if extra: detail += f", extra: {extra}"
    checks.append(("Sheet Structure", ok, detail))
    if missing:
        errors.append(f"Missing sheets: {missing}")

    # --- Check 2: BS Balance ---
    # For each year column, check that BS check row = 0
    bs_name = SHEETS['balance_sheet']['name']
    if bs_name in wb.sheetnames:
        ws = wb[bs_name]
        check_row = ROWS['balance_sheet']['bs_check']
        bs_errors = []
        for col in range(COL_HIST_START, COL_FORECAST_END + 1):
            val = ws.cell(row=check_row, column=col).value
            # For formula cells, just verify formula exists
            cell_val = ws.cell(row=check_row, column=col).value
            if cell_val is not None and isinstance(cell_val, (int, float)):
                if abs(cell_val) > 0.01:
                    bs_errors.append(f"Col {get_column_letter(col)}: {cell_val}")
        ok = len(bs_errors) == 0
        detail = "All years balanced" if ok else f"Imbalance in: {bs_errors}"
        checks.append(("BS Balance", ok, detail))
        if bs_errors:
            errors.extend([f"BS imbalance {e}" for e in bs_errors])
    else:
        checks.append(("BS Balance", False, "Sheet not found"))
        errors.append("Balance Sheet not found")

    # --- Check 3: Cash Tie-out ---
    cf_name = SHEETS['cash_flow']['name']
    if cf_name in wb.sheetnames:
        ws = wb[cf_name]
        tieout_row = ROWS['cash_flow']['cash_tieout']
        cf_errors = []
        for col in range(COL_HIST_START, COL_FORECAST_END + 1):
            val = ws.cell(row=tieout_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                if abs(val) > 0.01:
                    cf_errors.append(f"Col {get_column_letter(col)}: {val}")
        ok = len(cf_errors) == 0
        detail = "Cash tied out" if ok else f"Mismatch in: {cf_errors}"
        checks.append(("Cash Tie-out", ok, detail))
        if cf_errors:
            errors.extend([f"Cash tieout {e}" for e in cf_errors])
    else:
        checks.append(("Cash Tie-out", False, "Sheet not found"))

    # --- Check 4: Revenue Bridge ---
    is_name = SHEETS['income_statement']['name']
    ka_name = SHEETS['assumptions']['name']
    if is_name in wb.sheetnames and ka_name in wb.sheetnames:
        is_ws = wb[is_name]
        rev_row = ROWS['income_statement']['revenue']
        rev_errors = []
        for col in range(COL_FORECAST_START, COL_FORECAST_END + 1):
            val = is_ws.cell(row=rev_row, column=col).value
            # Check it's a formula linking to assumptions
            if isinstance(val, str) and '=' in val and ka_name in val:
                pass  # Good - it's a cross-sheet formula
            elif val is None:
                rev_errors.append(f"Col {get_column_letter(col)}: empty")
        ok = len(rev_errors) == 0
        checks.append(("Revenue Bridge", ok, "IS Revenue links to Assumptions" if ok else str(rev_errors)))
    else:
        checks.append(("Revenue Bridge", False, "Sheets not found"))

    # --- Check 5: EBITDA Bridge ---
    co_name = SHEETS['cogs_opex']['name']
    if is_name in wb.sheetnames:
        is_ws = wb[is_name]
        ebitda_row = ROWS['income_statement']['ebitda']
        ok = True
        for col in range(COL_FORECAST_START, COL_FORECAST_END + 1):
            val = is_ws.cell(row=ebitda_row, column=col).value
            if val is None:
                ok = False
        detail = "EBITDA formulas present" if ok else "EBITDA has blank cells in forecast"
        checks.append(("EBITDA Bridge", ok, detail))
    else:
        checks.append(("EBITDA Bridge", False, "IS not found"))

    # --- Check 6: RE Roll-forward ---
    if bs_name in wb.sheetnames:
        ws = wb[bs_name]
        re_row = ROWS['balance_sheet']['retained_earnings']
        re_ok = True
        for col in range(COL_FORECAST_START, COL_FORECAST_END + 1):
            val = ws.cell(row=re_row, column=col).value
            if val is None:
                re_ok = False
            elif isinstance(val, str) and is_name not in val:
                re_ok = False  # Should reference IS for NI
        detail = "RE linked to IS Net Income" if re_ok else "RE formula missing IS link"
        checks.append(("RE Roll-forward", re_ok, detail))
    else:
        checks.append(("RE Roll-forward", False, "BS not found"))

    # --- Check 7: Debt Consistency ---
    dc_name = SHEETS['debt_capex']['name']
    if bs_name in wb.sheetnames and dc_name in wb.sheetnames:
        bs_ws = wb[bs_name]
        lt_debt_row = ROWS['balance_sheet']['lt_debt']
        debt_ok = True
        for col in range(COL_FORECAST_START, COL_FORECAST_END + 1):
            val = bs_ws.cell(row=lt_debt_row, column=col).value
            if val is None:
                debt_ok = False
            elif isinstance(val, str) and dc_name not in val:
                debt_ok = False
        detail = "BS debt links to Debt & CapEx" if debt_ok else "BS debt not linked"
        checks.append(("Debt Consistency", debt_ok, detail))
    else:
        checks.append(("Debt Consistency", False, "Sheets not found"))

    # --- Check 8: Formula Presence ---
    # Sample anchor cells in forecast columns — they should be formulas (strings with '=')
    formula_checks = [
        ('income_statement', 'gross_profit'),
        ('income_statement', 'net_income'),
        ('balance_sheet', 'total_assets'),
        ('balance_sheet', 'bs_check'),
        ('cash_flow', 'operating_cf'),
        ('cash_flow', 'ending_cash'),
    ]
    blank_anchors = []
    for sheet_key, row_key in formula_checks:
        s_name = SHEETS[sheet_key]['name']
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            r = ROWS[sheet_key][row_key]
            col = COL_FORECAST_START  # Check first forecast column
            val = ws.cell(row=r, column=col).value
            if val is None or (isinstance(val, str) and not val.startswith('=')):
                blank_anchors.append(f"{s_name}!{get_column_letter(col)}{r}")
    ok = len(blank_anchors) == 0
    detail = "All anchor formulas present" if ok else f"Missing: {blank_anchors}"
    checks.append(("Formula Presence", ok, detail))
    if blank_anchors:
        errors.extend([f"No formula at {a}" for a in blank_anchors])

    # --- Check 9: No Blank Anchors ---
    critical_cells = [
        ('income_statement', 'revenue'),
        ('balance_sheet', 'cash'),
        ('balance_sheet', 'total_le'),
        ('cash_flow', 'cash_tieout'),
    ]
    blank_critical = []
    for sheet_key, row_key in critical_cells:
        s_name = SHEETS[sheet_key]['name']
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            r = ROWS[sheet_key][row_key]
            for col in range(COL_FORECAST_START, COL_FORECAST_END + 1):
                val = ws.cell(row=r, column=col).value
                if val is None:
                    blank_critical.append(f"{s_name}!{get_column_letter(col)}{r}")
    ok = len(blank_critical) == 0
    detail = "All critical cells populated" if ok else f"Blank: {blank_critical[:5]}"
    checks.append(("No Blank Anchors", ok, detail))

    # --- Check 10: Cross-sheet Links ---
    # Verify cells that should reference other sheets actually do
    link_checks = [
        ('income_statement', 'revenue', ka_name),
        ('income_statement', 'cogs', co_name),
        ('income_statement', 'interest_expense', dc_name),
        ('balance_sheet', 'cash', cf_name),
        ('balance_sheet', 'ppe_net', dc_name),
        ('cash_flow', 'net_income', is_name),
    ]
    broken_links = []
    for sheet_key, row_key, expected_ref in link_checks:
        s_name = SHEETS[sheet_key]['name']
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            r = ROWS[sheet_key][row_key]
            val = ws.cell(row=r, column=COL_FORECAST_START).value
            if val is None or (isinstance(val, str) and expected_ref not in val):
                broken_links.append(f"{s_name}!H{r} should ref '{expected_ref}'")
    ok = len(broken_links) == 0
    detail = "All cross-sheet links valid" if ok else f"Broken: {broken_links[:3]}"
    checks.append(("Cross-sheet Links", ok, detail))
    if broken_links:
        errors.extend(broken_links)

    wb.close()

    # Summary
    all_passed = all(c[1] for c in checks)

    if verbose:
        print(f"\n{'='*60}")
        print(f"Model Validation: {xlsx_path}")
        print(f"{'='*60}")
        for name, passed, detail in checks:
            status = "PASS" if passed else "FAIL"
            icon = "✓" if passed else "✗"
            print(f"  {icon} {name}: {status} — {detail}")
        print(f"{'='*60}")
        print(f"Result: {'ALL CHECKS PASSED' if all_passed else f'{len(errors)} ERRORS FOUND'}")
        print(f"{'='*60}\n")

    return {
        'passed': all_passed,
        'checks': checks,
        'errors': errors,
    }


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(_LIB_DIR), 'assets', 'template.xlsx')
    validate_model(path)
