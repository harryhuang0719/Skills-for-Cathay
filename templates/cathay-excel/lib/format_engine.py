"""
Cathay Capital PE Financial Model — Format Engine.

Applies Cathay brand formatting to Excel worksheets:
column widths, headers, fonts, borders, number formats, alternating rows,
and the historical/forecast divider.
"""

import sys
import os

_LIB_DIR = os.path.dirname(os.path.abspath(__file__))
if _LIB_DIR not in sys.path:
    sys.path.insert(0, _LIB_DIR)

from constants import *
from row_map import ROWS, SHEETS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter


# =============================================================================
# 1. Column Widths
# =============================================================================

def apply_column_widths(ws):
    """Set column widths from COL_WIDTHS constant."""
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# =============================================================================
# 2. Header & Subheader Rows
# =============================================================================

def apply_header_row(ws, row_num, text, merge_end_col=COL_LAST):
    """Cathay red background, white bold font, merged across columns."""
    ws.merge_cells(
        start_row=row_num, start_column=1,
        end_row=row_num, end_column=merge_end_col
    )
    cell = ws.cell(row=row_num, column=1)
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = ALIGN_HEADER
    # Apply fill to all cells in the merged range
    for col in range(2, merge_end_col + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT


def apply_subheader_row(ws, row_num, text):
    """Light pink background, dark red bold font, column A only."""
    cell = ws.cell(row=row_num, column=1)
    cell.value = text
    cell.font = SUBHEAD_FONT
    cell.fill = SUBHEAD_FILL
    cell.alignment = ALIGN_LEFT
    # Extend fill across data columns for visual consistency
    for col in range(2, COL_LAST + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = SUBHEAD_FILL


# =============================================================================
# 3. Year Row
# =============================================================================

def apply_year_row(ws, row_num):
    """Write year labels (2021-2029) in columns D-L, centered, bold.
    Thick right border on column G (hist/forecast divider).
    """
    for i, year in enumerate(ALL_YEARS):
        col = COL_HIST_START + i
        cell = ws.cell(row=row_num, column=col)
        # Add 'E' suffix for forecast years
        if year in FORECAST_YEARS:
            cell.value = f"{year}E"
        else:
            cell.value = year
        cell.font = LABEL_BOLD_FONT
        cell.alignment = ALIGN_CENTER
        cell.number_format = NUM_FMT_YEAR

    # Thick right border on hist/forecast divider (column G)
    divider_cell = ws.cell(row=row_num, column=COL_HIST_END)
    divider_cell.border = THICK_RIGHT_BORDER


# =============================================================================
# 4. Labels
# =============================================================================

def apply_label(ws, row_num, label, unit='', note='', bold=False, indent=0):
    """Write label in col A, unit in col B, note in col C.
    Indent by prepending spaces if indent > 0.
    """
    display_label = ('  ' * indent) + label if indent > 0 else label

    # Column A: label
    cell_a = ws.cell(row=row_num, column=COL_LABEL)
    cell_a.value = display_label
    cell_a.font = LABEL_BOLD_FONT if bold else LABEL_FONT
    cell_a.alignment = ALIGN_LEFT

    # Column B: unit
    if unit:
        cell_b = ws.cell(row=row_num, column=COL_UNIT)
        cell_b.value = unit
        cell_b.font = UNIT_FONT
        cell_b.alignment = ALIGN_CENTER

    # Column C: note
    if note:
        cell_c = ws.cell(row=row_num, column=COL_NOTE)
        cell_c.value = note
        cell_c.font = NOTE_FONT
        cell_c.alignment = ALIGN_LEFT


# =============================================================================
# 5. Total & Subtotal Row Formatting
# =============================================================================

def apply_total_row_format(ws, row_num):
    """Bold labels, top-thin + bottom-double border on data cells."""
    # Bold the label
    ws.cell(row=row_num, column=COL_LABEL).font = LABEL_BOLD_FONT
    # Apply border to data cells
    for col in range(COL_HIST_START, COL_LAST + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = LABEL_BOLD_FONT
        cell.border = BOTTOM_TOTAL_BORDER


def apply_subtotal_row_format(ws, row_num):
    """Bold labels, bottom-thin border on data cells."""
    ws.cell(row=row_num, column=COL_LABEL).font = LABEL_BOLD_FONT
    for col in range(COL_HIST_START, COL_LAST + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = LABEL_BOLD_FONT
        cell.border = BOTTOM_SUBTOTAL_BORDER


# =============================================================================
# 6. Check Row (conditional formatting)
# =============================================================================

def apply_check_row_format(ws, row_num):
    """Value=0 -> green PASS fill, else red fill.
    Applied via cell-level styling (check at generation time if values exist).
    """
    ws.cell(row=row_num, column=COL_LABEL).font = LABEL_BOLD_FONT
    for col in range(COL_HIST_START, COL_LAST + 1):
        cell = ws.cell(row=row_num, column=col)
        # Default to pass fill; template_builder or model_populator can override
        cell.fill = CHECK_PASS_FILL
        cell.font = LABEL_BOLD_FONT
        cell.alignment = ALIGN_CENTER


# =============================================================================
# 7. Alternating Rows
# =============================================================================

def apply_alternating_rows(ws, start_row, end_row):
    """Light grey on even rows (data columns only, D-L)."""
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for col in range(COL_HIST_START, COL_LAST + 1):
                ws.cell(row=r, column=col).fill = ALT_ROW_FILL


# =============================================================================
# 8. Number Formats
# =============================================================================

def apply_number_format(ws, row_num, fmt, start_col=COL_HIST_START, end_col=COL_LAST):
    """Set number format for a row's data cells."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).number_format = fmt


# =============================================================================
# 9. Historical/Forecast Divider
# =============================================================================

def apply_hist_forecast_divider(ws, start_row, end_row):
    """Thick right border on COL_HIST_END (G) for all rows."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=COL_HIST_END)
        # Preserve existing border properties, add thick right
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=Side(style='medium', color=CATHAY_BLACK),
            top=existing.top,
            bottom=existing.bottom,
        )


# =============================================================================
# 10. Font Application (Input / Formula / Link)
# =============================================================================

def apply_input_font(ws, row_num, start_col, end_col):
    """Blue font for hardcoded input cells."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).font = INPUT_FONT


def apply_formula_font(ws, row_num, start_col, end_col):
    """Black font for formula cells."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).font = FORMULA_FONT


def apply_link_font(ws, row_num, start_col, end_col):
    """Green font for cross-sheet links."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).font = LINK_FONT


# =============================================================================
# 11. Sheet-level Format Definitions
# =============================================================================

# Define which rows are totals, subtotals, percentages, linked, etc. per sheet

_SHEET_FORMAT_DEFS = {
    'assumptions': {
        'total_rows': ['total_revenue', 'total_sga_pct'],
        'subtotal_rows': ['seg_a_revenue', 'seg_b_revenue', 'seg_c_revenue'],
        'pct_rows': [
            'seg_a_growth', 'seg_a_pct', 'seg_b_growth', 'seg_b_pct',
            'seg_c_growth', 'seg_c_pct', 'total_growth',
            'cogs_pct_a', 'cogs_pct_b', 'cogs_pct_c', 'blended_cogs_pct', 'gross_margin',
            'personnel_pct', 'rent_pct', 'marketing_pct', 'rd_pct', 'other_opex_pct', 'total_sga_pct',
            'capex_pct', 'da_rate', 'tax_rate', 'dividend_payout', 'interest_rate',
        ],
        'subheader_rows': [
            'seg_a_header', 'seg_b_header', 'seg_c_header',
            'cost_header', 'opex_header', 'capex_header', 'wc_header', 'tax_header',
        ],
        'input_rows': [
            'seg_a_volume', 'seg_a_price', 'seg_a_utilization',
            'seg_b_volume', 'seg_b_price', 'seg_b_utilization',
            'seg_c_volume', 'seg_c_price', 'seg_c_utilization',
            'cogs_pct_a', 'cogs_pct_b', 'cogs_pct_c',
            'personnel_pct', 'rent_pct', 'marketing_pct', 'rd_pct', 'other_opex_pct',
            'capex_pct', 'da_rate', 'ar_days', 'ap_days', 'inventory_days',
            'tax_rate', 'dividend_payout', 'interest_rate',
        ],
        'link_rows': [],
    },
    'revenue': {
        'total_rows': ['bu_total', 'td_total'],
        'subtotal_rows': [],
        'pct_rows': ['penetration', 'market_share', 'recon_diff',
                     'mix_seg_a_pct', 'mix_seg_b_pct', 'mix_seg_c_pct',
                     'growth_total', 'growth_seg_a', 'growth_seg_b', 'growth_seg_c'],
        'subheader_rows': ['bu_header', 'td_header', 'recon_header', 'mix_header', 'growth_header'],
        'input_rows': ['bu_seg_a', 'bu_seg_b', 'bu_seg_c', 'tam'],
        'link_rows': [],
    },
    'cogs_opex': {
        'total_rows': ['total_cogs', 'gross_profit', 'total_sga', 'total_da', 'total_opex', 'ebitda', 'ebit'],
        'subtotal_rows': [],
        'pct_rows': ['cogs_pct_rev', 'gross_margin', 'sga_pct_rev', 'ebitda_margin', 'ebit_margin'],
        'subheader_rows': ['cogs_header', 'sga_header', 'da_header', 'ebitda_bridge_header'],
        'input_rows': [],
        'link_rows': ['cogs_seg_a', 'cogs_seg_b', 'cogs_seg_c',
                      'personnel', 'rent', 'marketing', 'rd', 'other_opex', 'depreciation'],
    },
    'income_statement': {
        'total_rows': ['gross_profit', 'ebitda', 'ebit', 'ebt', 'net_income'],
        'subtotal_rows': [],
        'pct_rows': ['gross_margin', 'ebitda_margin', 'ebit_margin', 'effective_tax_rate', 'net_margin'],
        'subheader_rows': [],
        'input_rows': ['other_income', 'other_fin', 'shares_outstanding'],
        'link_rows': ['revenue', 'cogs', 'sga', 'da', 'interest_expense', 'tax'],
    },
    'balance_sheet': {
        'total_rows': ['total_current_assets', 'total_noncurrent_assets', 'total_assets',
                       'total_current_liab', 'total_noncurrent_liab', 'total_liabilities',
                       'total_equity', 'total_le'],
        'subtotal_rows': [],
        'pct_rows': [],
        'subheader_rows': ['ca_header', 'nca_header', 'cl_header', 'ncl_header', 'eq_header'],
        'input_rows': ['other_current', 'intangibles', 'other_noncurrent',
                       'st_debt', 'other_current_liab', 'other_noncurrent_liab', 'paid_in_capital'],
        'link_rows': ['cash', 'accounts_receivable', 'inventory', 'ppe_net',
                      'accounts_payable', 'lt_debt', 'retained_earnings'],
        'check_rows': ['bs_check'],
    },
    'cash_flow': {
        'total_rows': ['operating_cf', 'investing_cf', 'financing_cf', 'net_cf', 'ending_cash'],
        'subtotal_rows': ['total_delta_wc'],
        'pct_rows': [],
        'subheader_rows': ['op_header', 'inv_header', 'fin_header'],
        'input_rows': ['delta_other_wc', 'other_investing', 'equity_issuance'],
        'link_rows': ['net_income', 'da', 'delta_ar', 'delta_inventory', 'delta_ap',
                      'capex', 'debt_drawdown', 'debt_repayment', 'dividends'],
        'check_rows': ['cash_tieout'],
    },
    'working_capital': {
        'total_rows': ['net_wc', 'delta_net_wc'],
        'subtotal_rows': [],
        'pct_rows': [],
        'subheader_rows': ['days_header', 'bal_header', 'change_header'],
        'input_rows': ['ar_days', 'inventory_days', 'ap_days'],
        'link_rows': ['ar_balance', 'inventory_balance', 'ap_balance'],
    },
    'debt_capex': {
        'total_rows': ['total_debt', 'total_interest', 'total_capex', 'ending_ppe'],
        'subtotal_rows': ['d1_ending', 'd2_ending', 'total_drawdown', 'total_repayment'],
        'pct_rows': ['d1_rate', 'd2_rate'],
        'subheader_rows': ['d1_header', 'd2_header', 'debt_summary_header', 'capex_header', 'da_header'],
        'input_rows': ['d1_drawdown', 'd1_repayment', 'd2_drawdown', 'd2_repayment',
                       'capex_maintenance', 'capex_growth'],
        'link_rows': [],
    },
    'returns': {
        'total_rows': ['ps_exit_equity', 'pe_exit_equity'],
        'subtotal_rows': ['ps_moic', 'ps_irr', 'pe_moic', 'pe_irr'],
        'pct_rows': ['ps_irr', 'pe_irr'],
        'subheader_rows': ['entry_header', 'ps_exit_header', 'pe_exit_header',
                           'sens1_header', 'sens2_header'],
        'input_rows': ['entry_revenue', 'entry_net_income', 'entry_ps_multiple', 'entry_pe_multiple',
                       'entry_ev', 'entry_net_debt', 'entry_equity',
                       'ps_exit_revenue', 'ps_exit_multiple', 'ps_exit_net_debt',
                       'pe_exit_ni', 'pe_exit_multiple'],
        'link_rows': [],
        'multiple_rows': ['ps_moic', 'pe_moic', 'entry_ps_multiple', 'entry_pe_multiple',
                          'ps_exit_multiple', 'pe_exit_multiple'],
    },
    'dcf': {
        'total_rows': ['ufcf', 'ev', 'equity_value'],
        'subtotal_rows': ['nopat', 'sum_pv_fcf', 'terminal_value', 'pv_tv'],
        'pct_rows': ['risk_free', 'mrp', 'cost_of_equity', 'cost_of_debt',
                     'tax_rate', 'equity_weight', 'debt_weight', 'wacc', 'terminal_growth'],
        'subheader_rows': ['wacc_header', 'ufcf_header', 'discount_header', 'tv_header', 'val_header'],
        'input_rows': ['risk_free', 'beta', 'mrp', 'cost_of_debt', 'tax_rate',
                       'equity_weight', 'debt_weight', 'terminal_growth', 'exit_multiple'],
        'link_rows': ['ebit', 'plus_da', 'less_capex', 'less_delta_wc'],
        'multiple_rows': ['exit_multiple', 'beta'],
    },
    'comps': {
        'total_rows': ['mean_row', 'median_row'],
        'subtotal_rows': ['q1_row', 'q3_row'],
        'pct_rows': [],
        'subheader_rows': ['implied_header'],
        'input_rows': ['comp_1'],  # All comp rows are inputs
        'link_rows': ['target_revenue', 'target_ebitda', 'target_ebit', 'target_ni'],
        'multiple_rows': ['implied_ev_rev', 'implied_ev_ebitda', 'implied_ev_ebit', 'implied_pe'],
    },
    'dashboard': {
        'total_rows': [],
        'subtotal_rows': [],
        'pct_rows': ['kpi_ebitda_margin', 'kpi_net_margin', 'kpi_revenue_cagr',
                     'return_ps_irr', 'return_pe_irr'],
        'subheader_rows': ['kpi_header', 'returns_header', 'check_header', 'metrics_header'],
        'input_rows': [],
        'link_rows': ['kpi_revenue', 'kpi_ebitda', 'kpi_ebitda_margin', 'kpi_net_income',
                      'kpi_net_margin', 'kpi_revenue_cagr',
                      'return_ps_irr', 'return_ps_moic', 'return_pe_irr', 'return_pe_moic',
                      'metrics_revenue', 'metrics_ebitda', 'metrics_ni', 'metrics_fcf'],
        'check_rows': ['bs_check_status', 'cash_check_status'],
        'multiple_rows': ['return_ps_moic', 'return_pe_moic'],
    },
}


# =============================================================================
# 12. Master Format Function
# =============================================================================

def format_sheet(ws, sheet_key):
    """Master function that applies all formatting for a given sheet.

    Applies:
    - Column widths
    - Header row (row 1)
    - Year row (row 3, if present)
    - Subheader rows
    - Total/subtotal row formatting
    - Number formats (pct, number, multiple)
    - Input font (blue) on historical columns for input rows
    - Formula font (black) on forecast columns for formula rows
    - Link font (green) on forecast columns for linked rows
    - Check row formatting
    - Hist/forecast divider
    - Alternating rows
    """
    r = ROWS.get(sheet_key)
    if r is None:
        return

    sheet_name = SHEETS[sheet_key]['name']
    fmt_def = _SHEET_FORMAT_DEFS.get(sheet_key, {})

    # 1. Column widths
    apply_column_widths(ws)

    # 2. Header row
    if 'header' in r:
        apply_header_row(ws, r['header'], sheet_name)

    # 3. Year row
    if 'year_row' in r:
        apply_year_row(ws, r['year_row'])

    # 4. Subheader rows
    for key in fmt_def.get('subheader_rows', []):
        if key in r:
            apply_subheader_row(ws, r[key], ws.cell(row=r[key], column=COL_LABEL).value or key.replace('_', ' ').title())

    # 5. Total rows
    for key in fmt_def.get('total_rows', []):
        if key in r:
            apply_total_row_format(ws, r[key])

    # 6. Subtotal rows
    for key in fmt_def.get('subtotal_rows', []):
        if key in r:
            apply_subtotal_row_format(ws, r[key])

    # 7. Percentage number format
    for key in fmt_def.get('pct_rows', []):
        if key in r:
            apply_number_format(ws, r[key], NUM_FMT_PCT)

    # 8. Multiple number format
    for key in fmt_def.get('multiple_rows', []):
        if key in r:
            apply_number_format(ws, r[key], NUM_FMT_MULTIPLE)

    # 9. Number format for non-pct, non-multiple data rows
    all_special = set(fmt_def.get('pct_rows', []) + fmt_def.get('multiple_rows', []))
    all_structural = set(
        fmt_def.get('subheader_rows', []) + ['header', 'year_row']
    )
    for key, row_num in r.items():
        if key not in all_special and key not in all_structural:
            # Apply standard number format to data rows
            apply_number_format(ws, row_num, NUM_FMT_NUMBER)

    # 10. Input font (blue) on historical columns
    for key in fmt_def.get('input_rows', []):
        if key in r:
            apply_input_font(ws, r[key], COL_HIST_START, COL_HIST_END)
            # Forecast columns for input rows also get input font (assumptions are inputs throughout)
            if sheet_key == 'assumptions':
                apply_input_font(ws, r[key], COL_FORECAST_START, COL_FORECAST_END)

    # 11. Link font (green) on forecast columns for linked rows
    for key in fmt_def.get('link_rows', []):
        if key in r:
            apply_link_font(ws, r[key], COL_FORECAST_START, COL_FORECAST_END)

    # 12. Formula font (black) on forecast columns for non-input, non-link rows
    input_set = set(fmt_def.get('input_rows', []))
    link_set = set(fmt_def.get('link_rows', []))
    for key, row_num in r.items():
        if key not in all_structural and key not in input_set and key not in link_set:
            apply_formula_font(ws, row_num, COL_FORECAST_START, COL_FORECAST_END)

    # 13. Check rows
    for key in fmt_def.get('check_rows', []):
        if key in r:
            apply_check_row_format(ws, r[key])

    # 14. Hist/forecast divider
    if 'year_row' in r:
        # Find the max row in this sheet
        max_row = max(r.values())
        apply_hist_forecast_divider(ws, r['year_row'], max_row)

    # 15. Alternating rows (skip header, year row, and subheaders)
    if 'year_row' in r:
        start = r['year_row'] + 1
        max_row = max(r.values())
        apply_alternating_rows(ws, start, max_row)
