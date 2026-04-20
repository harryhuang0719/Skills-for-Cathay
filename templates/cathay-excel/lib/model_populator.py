"""
Cathay Capital PE Financial Model — Model Populator.

Fills a blank template.xlsx with actual company data from a standardized data_dict.
Uses row_map for cell positioning and constants for styling.

Usage:
    from model_populator import populate_model
    populate_model("template.xlsx", data_dict, "output.xlsx")
"""

import sys
import os
import shutil

_LIB_DIR = os.path.dirname(os.path.abspath(__file__))
if _LIB_DIR not in sys.path:
    sys.path.insert(0, _LIB_DIR)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from constants import (
    COL_HIST_START, COL_HIST_END,
    COL_FORECAST_START, COL_FORECAST_END,
    HIST_YEARS, FORECAST_YEARS,
    INPUT_FONT, SEGMENT_LABELS,
)
from row_map import ROWS, SHEETS, row, cell_ref


# =============================================================================
# Public API
# =============================================================================

def populate_model(template_path, data_dict, output_path):
    """Fill a Cathay PE model template with company data.

    Args:
        template_path: path to blank template.xlsx
        data_dict: standardized dict with keys:
            'company_name': str
            'industry': str
            'date': str
            'analyst': str
            'currency': str (default 'RMB')
            'segments': list of 3 str (segment names)
            'historical': {
                2021: {'revenue': float, 'cogs': float, 'sga': float,
                       'da': float, 'interest': float, 'tax': float,
                       'cash': float, 'ar': float, 'inventory': float,
                       'ppe': float, 'ap': float, 'debt': float,
                       'equity': float, 'retained_earnings': float,
                       'capex': float, ...},
                2022: {...}, 2023: {...}, 2024: {...}
            },
            'assumptions': {
                'seg_a_name': str,
                'seg_b_name': str,
                'seg_c_name': str,
                'revenue_growth': [float x5],  # per forecast year
                'gross_margin_target': [float x5],
                'sga_pct': float,
                'capex_pct': float,
                'tax_rate': float,
                'ar_days': float,
                'ap_days': float,
                'inventory_days': float,
                'da_rate': float,
                'dividend_payout': float,
                'interest_rate': float,
            },
            'return_assumptions': {  # optional
                'entry_ps': float,
                'entry_pe': float,
                'exit_ps': float,
                'exit_pe': float,
                'exit_year': int,
            },
            'dcf_assumptions': {  # optional
                'risk_free': float,
                'beta': float,
                'mrp': float,
                'cost_of_debt': float,
                'equity_weight': float,
                'terminal_growth': float,
            },
        }
        output_path: where to save the populated model

    Returns:
        output_path
    """
    # Step 1: Copy template to output
    shutil.copy2(template_path, output_path)

    # Step 2: Open workbook
    wb = load_workbook(output_path)

    # Step 3: Fill each section
    _fill_cover(wb, data_dict)
    _fill_historical(wb, data_dict)
    _fill_assumptions(wb, data_dict)
    _rename_segments(wb, data_dict)

    if 'return_assumptions' in data_dict:
        _fill_return_assumptions(wb, data_dict['return_assumptions'])

    if 'dcf_assumptions' in data_dict:
        _fill_dcf_assumptions(wb, data_dict['dcf_assumptions'])

    # Step 7: Save
    wb.save(output_path)
    return output_path


# =============================================================================
# Internal: Cover Sheet
# =============================================================================

def _fill_cover(wb, data_dict):
    """Fill the Cover sheet with company metadata."""
    ws = wb[SHEETS['cover']['name']]

    cover_fields = {
        'company_name': data_dict.get('company_name', ''),
        'industry': data_dict.get('industry', ''),
        'date': data_dict.get('date', ''),
        'analyst': data_dict.get('analyst', ''),
        'base_currency': data_dict.get('currency', 'RMB'),
    }

    for field_key, value in cover_fields.items():
        r = row('cover', field_key)
        cell = ws.cell(row=r, column=COL_HIST_START)
        cell.value = value
        cell.font = INPUT_FONT


# =============================================================================
# Internal: Historical Data
# =============================================================================

def _fill_historical(wb, data_dict):
    """Fill historical financials into the Income Statement, Balance Sheet, etc."""
    historical = data_dict.get('historical', {})
    if not historical:
        return

    # Income Statement
    _fill_is_historical(wb, historical)

    # Balance Sheet
    _fill_bs_historical(wb, historical)

    # Cash Flow (capex)
    _fill_cf_historical(wb, historical)


def _fill_is_historical(wb, historical):
    """Fill Income Statement historical columns (D-G)."""
    ws = wb[SHEETS['income_statement']['name']]

    is_mapping = {
        'revenue': 'revenue',
        'cogs': 'cogs',
        'gross_profit': None,  # calculated
        'sga': 'sga',
        'da': 'da',
        'interest_expense': 'interest',
        'tax': 'tax',
        'net_income': None,  # calculated
    }

    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        if not year_data:
            continue

        for row_key, data_key in is_mapping.items():
            if data_key is None:
                continue
            value = year_data.get(data_key)
            if value is not None:
                r = row('income_statement', row_key)
                cell = ws.cell(row=r, column=col)
                cell.value = value
                cell.font = INPUT_FONT

        # Calculate and fill derived values
        revenue = year_data.get('revenue', 0)
        cogs = year_data.get('cogs', 0)
        sga = year_data.get('sga', 0)
        da = year_data.get('da', 0)
        interest = year_data.get('interest', 0)
        tax = year_data.get('tax', 0)

        gross_profit = revenue - cogs
        _set_input(ws, row('income_statement', 'gross_profit'), col, gross_profit)

        ebitda = gross_profit - sga
        _set_input(ws, row('income_statement', 'ebitda'), col, ebitda)

        ebit = ebitda - da
        _set_input(ws, row('income_statement', 'ebit'), col, ebit)

        ebt = ebit - interest
        _set_input(ws, row('income_statement', 'ebt'), col, ebt)

        net_income = ebt - tax
        _set_input(ws, row('income_statement', 'net_income'), col, net_income)

        # Margins (as decimals)
        if revenue:
            _set_input(ws, row('income_statement', 'gross_margin'), col, gross_profit / revenue)
            _set_input(ws, row('income_statement', 'ebitda_margin'), col, ebitda / revenue)
            _set_input(ws, row('income_statement', 'ebit_margin'), col, ebit / revenue)
            _set_input(ws, row('income_statement', 'net_margin'), col, net_income / revenue)


def _fill_bs_historical(wb, historical):
    """Fill Balance Sheet historical columns."""
    ws = wb[SHEETS['balance_sheet']['name']]

    bs_mapping = {
        'cash': 'cash',
        'accounts_receivable': 'ar',
        'inventory': 'inventory',
        'ppe_net': 'ppe',
        'accounts_payable': 'ap',
        'lt_debt': 'debt',
        'paid_in_capital': 'equity',
        'retained_earnings': 'retained_earnings',
    }

    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        if not year_data:
            continue

        for row_key, data_key in bs_mapping.items():
            value = year_data.get(data_key)
            if value is not None:
                r = row('balance_sheet', row_key)
                _set_input(ws, r, col, value)

        # Calculate totals
        cash = year_data.get('cash', 0)
        ar = year_data.get('ar', 0)
        inventory = year_data.get('inventory', 0)
        other_ca = year_data.get('other_current_assets', 0)
        total_ca = cash + ar + inventory + other_ca
        _set_input(ws, row('balance_sheet', 'total_current_assets'), col, total_ca)

        ppe = year_data.get('ppe', 0)
        intangibles = year_data.get('intangibles', 0)
        other_nca = year_data.get('other_noncurrent_assets', 0)
        total_nca = ppe + intangibles + other_nca
        _set_input(ws, row('balance_sheet', 'total_noncurrent_assets'), col, total_nca)

        total_assets = total_ca + total_nca
        _set_input(ws, row('balance_sheet', 'total_assets'), col, total_assets)

        ap = year_data.get('ap', 0)
        st_debt = year_data.get('st_debt', 0)
        other_cl = year_data.get('other_current_liab', 0)
        total_cl = ap + st_debt + other_cl
        _set_input(ws, row('balance_sheet', 'total_current_liab'), col, total_cl)

        lt_debt = year_data.get('debt', 0)
        other_ncl = year_data.get('other_noncurrent_liab', 0)
        total_ncl = lt_debt + other_ncl
        _set_input(ws, row('balance_sheet', 'total_noncurrent_liab'), col, total_ncl)

        total_liab = total_cl + total_ncl
        _set_input(ws, row('balance_sheet', 'total_liabilities'), col, total_liab)

        equity = year_data.get('equity', 0)
        retained = year_data.get('retained_earnings', 0)
        total_equity = equity + retained
        _set_input(ws, row('balance_sheet', 'total_equity'), col, total_equity)

        total_le = total_liab + total_equity
        _set_input(ws, row('balance_sheet', 'total_le'), col, total_le)

        # BS check
        _set_input(ws, row('balance_sheet', 'bs_check'), col, total_assets - total_le)


def _fill_cf_historical(wb, historical):
    """Fill Cash Flow historical columns (capex and basic items)."""
    ws = wb[SHEETS['cash_flow']['name']]

    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        if not year_data:
            continue

        capex = year_data.get('capex')
        if capex is not None:
            _set_input(ws, row('cash_flow', 'capex'), col, -abs(capex))

    # Also fill the Debt & CapEx sheet
    ws_dc = wb[SHEETS['debt_capex']['name']]
    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        if not year_data:
            continue

        capex = year_data.get('capex')
        if capex is not None:
            _set_input(ws_dc, row('debt_capex', 'total_capex'), col, abs(capex))

        ppe = year_data.get('ppe')
        if ppe is not None:
            _set_input(ws_dc, row('debt_capex', 'ending_ppe'), col, ppe)


# =============================================================================
# Internal: Assumptions
# =============================================================================

def _fill_assumptions(wb, data_dict):
    """Fill the Key Assumptions sheet with forecast drivers."""
    assumptions = data_dict.get('assumptions', {})
    if not assumptions:
        return

    ws = wb[SHEETS['assumptions']['name']]
    historical = data_dict.get('historical', {})

    # Revenue growth rates (5 forecast years)
    revenue_growth = assumptions.get('revenue_growth', [])
    for i, growth in enumerate(revenue_growth[:5]):
        col = COL_FORECAST_START + i
        _set_input(ws, row('assumptions', 'total_growth'), col, growth)

    # Compute historical segment splits for revenue buildup
    _fill_segment_assumptions(ws, historical, assumptions)

    # Gross margin targets -> COGS percentages
    gross_margin_target = assumptions.get('gross_margin_target', [])
    for i, gm in enumerate(gross_margin_target[:5]):
        col = COL_FORECAST_START + i
        _set_input(ws, row('assumptions', 'gross_margin'), col, gm)
        # Implied blended COGS = 1 - gross_margin
        _set_input(ws, row('assumptions', 'blended_cogs_pct'), col, 1.0 - gm)

    # SG&A breakdown (flat assumption across forecast)
    sga_pct = assumptions.get('sga_pct', 0)
    personnel_pct = assumptions.get('personnel_pct', sga_pct * 0.5)
    rent_pct = assumptions.get('rent_pct', sga_pct * 0.1)
    marketing_pct = assumptions.get('marketing_pct', sga_pct * 0.15)
    rd_pct = assumptions.get('rd_pct', sga_pct * 0.15)
    other_opex_pct = assumptions.get('other_opex_pct', sga_pct * 0.1)

    opex_items = {
        'personnel_pct': personnel_pct,
        'rent_pct': rent_pct,
        'marketing_pct': marketing_pct,
        'rd_pct': rd_pct,
        'other_opex_pct': other_opex_pct,
        'total_sga_pct': sga_pct,
    }

    for row_key, value in opex_items.items():
        for i in range(5):
            col = COL_FORECAST_START + i
            _set_input(ws, row('assumptions', row_key), col, value)

    # CapEx & D&A
    capex_pct = assumptions.get('capex_pct', 0)
    da_rate = assumptions.get('da_rate', 0)
    for i in range(5):
        col = COL_FORECAST_START + i
        _set_input(ws, row('assumptions', 'capex_pct'), col, capex_pct)
        _set_input(ws, row('assumptions', 'da_rate'), col, da_rate)

    # Working capital days
    ar_days = assumptions.get('ar_days', 0)
    ap_days = assumptions.get('ap_days', 0)
    inventory_days = assumptions.get('inventory_days', 0)
    for i in range(5):
        col = COL_FORECAST_START + i
        _set_input(ws, row('assumptions', 'ar_days'), col, ar_days)
        _set_input(ws, row('assumptions', 'ap_days'), col, ap_days)
        _set_input(ws, row('assumptions', 'inventory_days'), col, inventory_days)

    # Tax & other
    tax_rate = assumptions.get('tax_rate', 0)
    dividend_payout = assumptions.get('dividend_payout', 0)
    interest_rate = assumptions.get('interest_rate', 0)
    for i in range(5):
        col = COL_FORECAST_START + i
        _set_input(ws, row('assumptions', 'tax_rate'), col, tax_rate)
        _set_input(ws, row('assumptions', 'dividend_payout'), col, dividend_payout)
        _set_input(ws, row('assumptions', 'interest_rate'), col, interest_rate)

    # Also fill historical assumptions (derived from actual data)
    _fill_historical_assumptions(ws, historical)


def _fill_segment_assumptions(ws, historical, assumptions):
    """Fill segment volume/price/growth from historical base + assumptions."""
    # Get last historical year as base
    last_year = max(historical.keys()) if historical else 2024
    base_data = historical.get(last_year, {})
    base_revenue = base_data.get('revenue', 0)

    # Segment revenue splits (from historical or equal split)
    seg_splits = assumptions.get('segment_splits', [0.4, 0.35, 0.25])

    # Fill segment revenue for historical years
    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        rev = year_data.get('revenue', 0)

        # Use actual segment data if available, else apply splits
        seg_a_rev = year_data.get('seg_a_revenue', rev * seg_splits[0])
        seg_b_rev = year_data.get('seg_b_revenue', rev * seg_splits[1])
        seg_c_rev = year_data.get('seg_c_revenue', rev * seg_splits[2])

        _set_input(ws, row('assumptions', 'seg_a_revenue'), col, seg_a_rev)
        _set_input(ws, row('assumptions', 'seg_b_revenue'), col, seg_b_rev)
        _set_input(ws, row('assumptions', 'seg_c_revenue'), col, seg_c_rev)
        _set_input(ws, row('assumptions', 'total_revenue'), col, rev)

        # Segment percentages
        if rev:
            _set_input(ws, row('assumptions', 'seg_a_pct'), col, seg_a_rev / rev)
            _set_input(ws, row('assumptions', 'seg_b_pct'), col, seg_b_rev / rev)
            _set_input(ws, row('assumptions', 'seg_c_pct'), col, seg_c_rev / rev)

    # Fill forecast segment revenue using growth rates
    revenue_growth = assumptions.get('revenue_growth', [0.1] * 5)
    prev_rev = base_revenue
    for i, growth in enumerate(revenue_growth[:5]):
        col = COL_FORECAST_START + i
        forecast_rev = prev_rev * (1 + growth)

        seg_a_rev = forecast_rev * seg_splits[0]
        seg_b_rev = forecast_rev * seg_splits[1]
        seg_c_rev = forecast_rev * seg_splits[2]

        _set_input(ws, row('assumptions', 'seg_a_revenue'), col, seg_a_rev)
        _set_input(ws, row('assumptions', 'seg_b_revenue'), col, seg_b_rev)
        _set_input(ws, row('assumptions', 'seg_c_revenue'), col, seg_c_rev)
        _set_input(ws, row('assumptions', 'total_revenue'), col, forecast_rev)
        _set_input(ws, row('assumptions', 'total_growth'), col, growth)

        # Segment percentages (stable splits)
        _set_input(ws, row('assumptions', 'seg_a_pct'), col, seg_splits[0])
        _set_input(ws, row('assumptions', 'seg_b_pct'), col, seg_splits[1])
        _set_input(ws, row('assumptions', 'seg_c_pct'), col, seg_splits[2])

        prev_rev = forecast_rev


def _fill_historical_assumptions(ws, historical):
    """Derive assumption values from historical data and fill them."""
    for year_idx, year in enumerate(HIST_YEARS):
        col = COL_HIST_START + year_idx
        year_data = historical.get(year, {})
        if not year_data:
            continue

        revenue = year_data.get('revenue', 0)
        cogs = year_data.get('cogs', 0)
        sga = year_data.get('sga', 0)

        if revenue:
            # Blended COGS %
            _set_input(ws, row('assumptions', 'blended_cogs_pct'), col, cogs / revenue)
            _set_input(ws, row('assumptions', 'gross_margin'), col, 1 - cogs / revenue)
            # SG&A %
            _set_input(ws, row('assumptions', 'total_sga_pct'), col, sga / revenue)

        # Growth (YoY)
        if year_idx > 0:
            prev_year_data = historical.get(HIST_YEARS[year_idx - 1], {})
            prev_rev = prev_year_data.get('revenue', 0)
            if prev_rev:
                _set_input(ws, row('assumptions', 'total_growth'), col, (revenue - prev_rev) / prev_rev)

        # Working capital days
        ar = year_data.get('ar', 0)
        ap = year_data.get('ap', 0)
        inventory = year_data.get('inventory', 0)

        if revenue:
            _set_input(ws, row('assumptions', 'ar_days'), col, ar / revenue * 365)
        if cogs:
            _set_input(ws, row('assumptions', 'inventory_days'), col, inventory / cogs * 365)
            _set_input(ws, row('assumptions', 'ap_days'), col, ap / cogs * 365)

        # Tax rate
        ebt = revenue - cogs - sga - year_data.get('da', 0) - year_data.get('interest', 0)
        tax = year_data.get('tax', 0)
        if ebt > 0:
            _set_input(ws, row('assumptions', 'tax_rate'), col, tax / ebt)

        # CapEx % of revenue
        capex = year_data.get('capex', 0)
        if revenue:
            _set_input(ws, row('assumptions', 'capex_pct'), col, capex / revenue)

        # D&A rate (% of PPE)
        ppe = year_data.get('ppe', 0)
        da = year_data.get('da', 0)
        if ppe:
            _set_input(ws, row('assumptions', 'da_rate'), col, da / ppe)


# =============================================================================
# Internal: Return Assumptions
# =============================================================================

def _fill_return_assumptions(wb, return_assumptions):
    """Fill Returns & Sensitivity sheet entry/exit assumptions."""
    ws = wb[SHEETS['returns']['name']]

    # Entry multiples (placed in the entry year column)
    entry_col = COL_HIST_END  # column G = 2024 = entry year

    entry_ps = return_assumptions.get('entry_ps')
    if entry_ps is not None:
        _set_input(ws, row('returns', 'entry_ps_multiple'), entry_col, entry_ps)

    entry_pe = return_assumptions.get('entry_pe')
    if entry_pe is not None:
        _set_input(ws, row('returns', 'entry_pe_multiple'), entry_col, entry_pe)

    # Exit multiples (placed in exit year column)
    exit_year = return_assumptions.get('exit_year', 2029)
    exit_col = COL_FORECAST_START + (exit_year - FORECAST_YEARS[0])
    exit_col = min(exit_col, COL_FORECAST_END)

    exit_ps = return_assumptions.get('exit_ps')
    if exit_ps is not None:
        _set_input(ws, row('returns', 'ps_exit_multiple'), exit_col, exit_ps)

    exit_pe = return_assumptions.get('exit_pe')
    if exit_pe is not None:
        _set_input(ws, row('returns', 'pe_exit_multiple'), exit_col, exit_pe)


# =============================================================================
# Internal: DCF Assumptions
# =============================================================================

def _fill_dcf_assumptions(wb, dcf_assumptions):
    """Fill DCF Valuation sheet WACC components."""
    ws = wb[SHEETS['dcf']['name']]

    # WACC inputs are in a fixed column (D = COL_HIST_START)
    val_col = COL_HIST_START

    dcf_mapping = {
        'risk_free': 'risk_free',
        'beta': 'beta',
        'mrp': 'mrp',
        'cost_of_debt': 'cost_of_debt',
        'equity_weight': 'equity_weight',
        'terminal_growth': 'terminal_growth',
    }

    for row_key, data_key in dcf_mapping.items():
        value = dcf_assumptions.get(data_key)
        if value is not None:
            _set_input(ws, row('dcf', row_key), val_col, value)

    # Debt weight = 1 - equity weight
    equity_weight = dcf_assumptions.get('equity_weight')
    if equity_weight is not None:
        _set_input(ws, row('dcf', 'debt_weight'), val_col, 1.0 - equity_weight)


# =============================================================================
# Internal: Segment Renaming
# =============================================================================

def _rename_segments(wb, data_dict):
    """Replace default segment labels with actual segment names."""
    segments = data_dict.get('segments', [])
    if not segments or len(segments) < 3:
        return

    # Build replacement map
    replacements = {
        'Segment A': segments[0],
        'Segment B': segments[1],
        'Segment C': segments[2],
        'SEGMENT A': segments[0].upper(),
        'SEGMENT B': segments[1].upper(),
        'SEGMENT C': segments[2].upper(),
    }

    # Scan label columns (column A) in all sheets
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=1)  # Column A = labels
            if cell.value and isinstance(cell.value, str):
                for old, new in replacements.items():
                    if old in cell.value:
                        cell.value = cell.value.replace(old, new)
                        break


# =============================================================================
# Helpers
# =============================================================================

def _set_input(ws, r, col, value):
    """Set a cell value with INPUT_FONT (blue, hardcoded input)."""
    cell = ws.cell(row=r, column=col)
    cell.value = value
    cell.font = INPUT_FONT
