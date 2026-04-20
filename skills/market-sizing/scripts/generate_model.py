#!/usr/bin/env python3
"""
Market Sizing + Supply-Demand + Competitive Landscape Excel Model Generator (v2)

Accepts a JSON config file defining:
  - Multi-level demand hierarchy (segments -> sub-segments -> end customers)
  - Supply players with listed/non-listed status, source quality, and notes

Generates a 6-sheet Excel workbook with full formula linkages.

Usage:
  python generate_model.py --config config.json --output output.xlsx
  python generate_model.py --config config.json  # defaults to /mnt/user-data/outputs/
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ── Style Constants ──────────────────────────────────────────────────────────

FONT_TITLE = Font(name='Arial', bold=True, size=14, color='1F3864')
FONT_SUBTITLE = Font(name='Arial', size=10, color='666666', italic=True)
FONT_HEADER = Font(name='Arial', bold=True, size=10, color='FFFFFF')
FONT_SECTION = Font(name='Arial', bold=True, size=11, color='1F3864')
FONT_L1 = Font(name='Arial', bold=True, size=10, color='000000')
FONT_L2 = Font(name='Arial', size=10, color='333333')
FONT_L3 = Font(name='Arial', size=9, color='666666')
FONT_INPUT = Font(name='Arial', size=10, color='0000FF')
FONT_INPUT_L2 = Font(name='Arial', size=9, color='0000FF')
FONT_FORMULA = Font(name='Arial', size=10, color='000000')
FONT_XREF = Font(name='Arial', size=10, color='008000')
FONT_LABEL = Font(name='Arial', size=10, color='000000')
FONT_CHECK = Font(name='Arial', size=10, color='FF0000', bold=True)
FONT_NOTE = Font(name='Arial', size=9, color='888888', italic=True)
FONT_SRC_A = Font(name='Arial', size=9, color='008000', bold=True)
FONT_SRC_B = Font(name='Arial', size=9, color='0000FF')
FONT_SRC_C = Font(name='Arial', size=9, color='FF8800')
FONT_SRC_D = Font(name='Arial', size=9, color='FF0000')

FILL_HEADER = PatternFill('solid', fgColor='1F3864')
FILL_SECTION = PatternFill('solid', fgColor='D6E4F0')
FILL_SUBSECTION = PatternFill('solid', fgColor='E8EFF7')
FILL_INPUT_KEY = PatternFill('solid', fgColor='FFFF00')
FILL_CHECK = PatternFill('solid', fgColor='FFF2CC')
FILL_PLAYER_LISTED = PatternFill('solid', fgColor='E2EFDA')
FILL_PLAYER_PRIVATE = PatternFill('solid', fgColor='FCE4D6')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

FMT_COMMA = '#,##0'
FMT_COMMA1 = '#,##0.0'
FMT_MONEY = '$#,##0'
FMT_PCT = '0.0%'

SOURCE_FONTS = {'A': FONT_SRC_A, 'B': FONT_SRC_B, 'C': FONT_SRC_C, 'D': FONT_SRC_D}


def col(n):
    return get_column_letter(n)


def cr(c, r):
    return f'{col(c)}{r}'


# ── Helpers ──────────────────────────────────────────────────────────────────

def hdr_row(ws, row, labels, start=1):
    for i, lb in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=lb)
        c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, ALIGN_CENTER, THIN_BORDER


def section_row(ws, row, label, ncols, start=1, fill=FILL_SECTION):
    ws.cell(row=row, column=start, value=label).font = FONT_SECTION
    for j in range(start, start + ncols):
        ws.cell(row=row, column=j).fill = fill
        ws.cell(row=row, column=j).border = THIN_BORDER


def w(ws, row, column, value, font=FONT_LABEL, fmt=None, fill=None, align=ALIGN_LEFT):
    c = ws.cell(row=row, column=column, value=value)
    c.font, c.border, c.alignment = font, THIN_BORDER, align
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c


def wi(ws, row, column, value=0, fmt=FMT_COMMA, font=FONT_INPUT, hl=False):
    c = ws.cell(row=row, column=column, value=value)
    c.font, c.border, c.alignment = font, THIN_BORDER, ALIGN_RIGHT
    if fmt: c.number_format = fmt
    if hl: c.fill = FILL_INPUT_KEY
    return c


def _get_seg_vols(seg, yn):
    """从 segment 数据提取各年 volume（用于 cached value 计算）"""
    subs = seg.get('sub_segments', [])
    if subs:
        totals = [0.0] * yn
        for sub in subs:
            vols = sub.get('volumes', []) if isinstance(sub, dict) else []
            for j in range(min(yn, len(vols))):
                totals[j] += vols[j] or 0
        return totals
    return seg.get('volumes', [0] * yn)


def wf(ws, row, column, formula, fmt=FMT_COMMA, font=FONT_FORMULA, cached=None):
    c = ws.cell(row=row, column=column, value=formula)
    c.font, c.border, c.alignment = font, THIN_BORDER, ALIGN_RIGHT
    if fmt: c.number_format = fmt
    return c


def set_widths(ws, widths):
    for i, ww in enumerate(widths, 1):
        ws.column_dimensions[col(i)].width = ww


def yoy(ws, r, ref_row, yr_n, label='  YoY Growth', font=FONT_L3):
    w(ws, r, 1, label, font)
    w(ws, r, 2, 'n/a', FONT_NOTE)
    for j in range(1, yr_n):
        prev, curr = cr(j + 1, ref_row), cr(j + 2, ref_row)
        wf(ws, r, j + 2, f'=IF({prev}=0,0,({curr}-{prev})/{prev})', FMT_PCT)


# ── ASSUMPTIONS ──────────────────────────────────────────────────────────────

def build_assumptions(wb, cfg):
    ws = wb.create_sheet('Assumptions', 0)
    years = cfg['years']; yn = len(years)
    nc = yn + 2
    set_widths(ws, [42] + [15] * yn + [55])

    r = 1; w(ws, r, 1, f"{cfg['title']} — Key Assumptions", FONT_TITLE)
    r = 2; w(ws, r, 1, 'Blue = editable | Yellow = key assumption | Right col = source', FONT_SUBTITLE)
    # P1-6: Unit definition
    r = 3
    unit_def = cfg.get('unit_definition') or f"1 {cfg.get('unit','Unit')} = (see market_boundary)"
    w(ws, r, 1, f"📐 Unit Definition: {cfg.get('unit','Unit')} | {unit_def}", FONT_SUBTITLE)
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='E8F4FD')
    r = 4; hdr_row(ws, r, ['Assumption'] + years + ['Source / Notes']); r += 1

    # ── v3.0: Objective Classifier + Archetype Router ──────────────────────
    section_row(ws, r, '🎯 OBJECTIVE CLASSIFIER (v3.0 — fill before modeling)', nc,
                fill=PatternFill('solid', fgColor='FFF2CC')); r += 1
    oc = cfg.get('objective_classifier', {})
    arch = cfg.get('primary_archetype', '')
    fc = cfg.get('formula_contract', {})
    oc_fields = [
        ('Sizing Objective',    oc.get('sizing_objective', ''),    'TAM / SAM / SOM / Current Market Size / Future Realizable Pool'),
        ('Realization Basis',   oc.get('realization_basis', ''),   'theoretical / serviceable / realizable / realized'),
        ('Time Basis',          oc.get('time_basis', ''),          'current / historical / forecast ≤3Y / forecast >3Y'),
        ('Value Basis',         oc.get('value_basis', ''),         'revenue / units / both'),
        ('Primary Archetype',   arch,                              'population / installed_base / throughput / project / substitution / commodity'),
        ('Secondary Archetype', cfg.get('secondary_archetype', 'None'), 'optional bridge archetype'),
        ('Generator Formula',   fc.get('volume_driver', ''),       'e.g. installed_base × attach_rate × replacement_rate'),
        ('ASP Mechanism',       fc.get('price_driver', ''),        'exogenous OR gap-driven (ASP_t = max(floor, ASP_{t-1}×(1+e×gap%)))'),
        ('Competition Mode',    fc.get('competition_driver', ''),  'min(supply, allocated_demand) × player_ASP'),
        ('Generator Bridge >3Y Required', str(cfg.get('generator_bridge_required', '')), 'true / false'),
        ('Why Primary Archetype', cfg.get('archetype_rationale', ''), 'explain why this archetype, not others'),
    ]
    for lb, val, src in oc_fields:
        w(ws, r, 1, f'  {lb}', FONT_L2)
        c = ws.cell(row=r, column=2, value=val or '')
        c.font = FONT_INPUT; c.fill = FILL_INPUT_KEY if not val else PatternFill('solid', fgColor='FFFACD')
        c.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=yn + 1)
        w(ws, r, yn + 2, src, FONT_NOTE)
        r += 1
    r += 1

    # ── Demand ──
    section_row(ws, r, 'DEMAND ASSUMPTIONS', nc); r += 1
    dm = {}  # demand_map

    # ── V3.3: Archetype-Driven Demand Engine ─────────────────────────────
    de = cfg.get('demand_engine')
    if de:
        params = de.get('params', {})
        section_row(ws, r, f'🔧 DEMAND ENGINE — {de.get("archetype", "unknown").upper()}', nc,
                    fill=PatternFill('solid', fgColor='E8F5E9')); r += 1
        w(ws, r, 1, f'  Formula: {de.get("formula", "N/A")}', FONT_NOTE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=yn + 1); r += 1

        param_rows = {}
        for pname, pvals in params.items():
            pr = r; param_rows[pname] = pr
            label = pname.replace('_', ' ').title()
            fmt = FMT_PCT if 'pct' in pname or 'rate' in pname else FMT_COMMA1
            w(ws, r, 1, f'  ⚙ {label}', FONT_L2)
            for j in range(yn):
                wi(ws, r, j + 2, pvals[j] if j < len(pvals) else 0, fmt, FONT_INPUT, hl=True)
            w(ws, r, yn + 2, f'demand_engine.params.{pname}', FONT_NOTE)
            r += 1

        # Computed total volume formula row
        engine_vol_row = r
        w(ws, r, 1, f'  ⚡ Engine Computed Volume ({cfg["unit"]})', FONT_L1)
        # Build formula from archetype: base × rate × (1/cycle) × 1000
        arch = de.get('archetype', '')
        for j in range(yn):
            col_j = j + 2
            if arch == 'installed_base' and all(k in param_rows for k in ('compatible_base_M', 'attach_rate_pct', 'replacement_cycle_yrs')):
                base_ref = cr(col_j, param_rows['compatible_base_M'])
                attach_ref = cr(col_j, param_rows['attach_rate_pct'])
                cycle_ref = cr(col_j, param_rows['replacement_cycle_yrs'])
                wf(ws, r, col_j, f'={base_ref}*{attach_ref}*(1/{cycle_ref})*1000', FMT_COMMA)
            elif arch == 'population' and all(k in param_rows for k in ('population_M', 'penetration_pct')):
                pop_ref = cr(col_j, param_rows['population_M'])
                pen_ref = cr(col_j, param_rows['penetration_pct'])
                wf(ws, r, col_j, f'={pop_ref}*{pen_ref}*1000', FMT_COMMA)
            elif arch == 'throughput' and all(k in param_rows for k in ('capacity_units', 'utilization_pct')):
                cap_ref = cr(col_j, param_rows['capacity_units'])
                util_ref = cr(col_j, param_rows['utilization_pct'])
                wf(ws, r, col_j, f'={cap_ref}*{util_ref}', FMT_COMMA)
            else:
                # Generic: multiply all param refs
                refs = [cr(col_j, param_rows[k]) for k in param_rows]
                wf(ws, r, col_j, '=' + '*'.join(refs), FMT_COMMA)
        dm['_engine_vol_row'] = engine_vol_row
        r += 1

        # Segment allocation % header
        section_row(ws, r, 'SEGMENT ALLOCATION (% of Engine Total)', nc,
                    fill=PatternFill('solid', fgColor='FFF8E1')); r += 1

    for seg in cfg['demand']['segments']:
        sn = seg['name']; dm[sn] = {}; subs = seg.get('sub_segments', [])

        if de:
            # ── V3.3 Engine mode: allocation % → computed volume ──
            alloc = seg.get('allocation_pct', [])
            alloc_row = r
            w(ws, r, 1, f'▸ {sn} — Allocation %', FONT_L2)
            for j in range(yn): wi(ws, r, j + 2, alloc[j] if j < len(alloc) else 0, FMT_PCT, FONT_INPUT, hl=True)
            w(ws, r, yn + 2, 'Share of engine total', FONT_NOTE)
            dm[sn]['_alloc_row'] = alloc_row; r += 1

            tvr = r
            w(ws, r, 1, f'    ↳ {sn} Computed Volume ({cfg["unit"]})', FONT_L1)
            evr = dm['_engine_vol_row']
            for j in range(yn):
                wf(ws, r, j + 2, f'={cr(j+2, evr)}*{cr(j+2, alloc_row)}', FMT_COMMA)
            dm[sn]['_total_vol_row'] = tvr; r += 1

        elif subs:
            w(ws, r, 1, f'▸ {sn}', FONT_L1)
            for jj in range(1, nc + 1): ws.cell(row=r, column=jj).fill = FILL_SUBSECTION
            r += 1

            sub_vols = []
            for sub in subs:
                sub_name = sub['name'] if isinstance(sub, dict) else sub
                vr = r; sub_vols.append(vr)
                w(ws, r, 1, f'    {sub_name} — Volume ({cfg["unit"]})', FONT_L2)
                vols = sub.get('volumes', []) if isinstance(sub, dict) else []
                for j in range(yn): wi(ws, r, j + 2, vols[j] if j < len(vols) else 0, FMT_COMMA, FONT_INPUT_L2, hl=True)
                src = sub.get('source', 'TBD') if isinstance(sub, dict) else 'TBD'
                w(ws, r, yn + 2, src, FONT_NOTE)
                dm[sn][sub_name] = {'vol_row': vr}
                r += 1

            tvr = r
            w(ws, r, 1, f'    ↳ {sn} Total Volume', FONT_L1)
            for j in range(yn):
                refs = '+'.join([cr(j + 2, v) for v in sub_vols])
                wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
            dm[sn]['_total_vol_row'] = tvr; r += 1
        else:
            vr = r
            vols = seg.get('volumes', [])
            w(ws, r, 1, f'▸ {sn} — Volume ({cfg["unit"]})', FONT_L1)
            for j in range(yn): wi(ws, r, j + 2, vols[j] if j < len(vols) else 0, FMT_COMMA, hl=True)
            w(ws, r, yn + 2, 'TBD', FONT_NOTE)
            dm[sn] = {'_total_vol_row': vr}; r += 1

        ar = r
        seg_asp = seg.get('asp_estimates', cfg.get('demand', {}).get('asp_estimates', []))
        w(ws, r, 1, f'    {sn} — ASP ({cfg.get("asp_label", "$")})', FONT_L2)
        for j in range(yn): wi(ws, r, j + 2, seg_asp[j] if j < len(seg_asp) else 0, FMT_MONEY, hl=True)
        dm[sn]['_asp_row'] = ar; r += 1; r += 1

    gvr = r
    w(ws, r, 1, f'TOTAL DEMAND VOLUME ({cfg["unit"]})', FONT_L1)
    vrefs = [dm[s['name']]['_total_vol_row'] for s in cfg['demand']['segments']]
    for j in range(yn):
        refs = '+'.join([cr(j + 2, v) for v in vrefs])
        wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
    r += 1

    # V3.3: Engine vs Bottom-Up cross-check
    if de:
        w(ws, r, 1, '  ✓ CHECK: Engine Total vs Σ Segments', FONT_CHECK)
        evr = dm['_engine_vol_row']
        for j in range(yn):
            wf(ws, r, j + 2, f'=IF({cr(j+2, evr)}=0,0,({cr(j+2, gvr)}-{cr(j+2, evr)})/{cr(j+2, evr)})', FMT_PCT, FONT_CHECK)
            ws.cell(row=r, column=j + 2).fill = FILL_CHECK
        r += 1
    r += 1

    # ── Supply ──
    section_row(ws, r, 'SUPPLY ASSUMPTIONS', nc); r += 1
    sm = {}

    for player in cfg['supply']['players']:
        pn = player['name']; il = player.get('listed', False)
        sq = player.get('source_quality', 'C'); notes = player.get('notes', '')
        badge = f"[{player.get('ticker', 'PRIVATE')}] [Src:{sq}]"
        pf = FILL_PLAYER_LISTED if il else FILL_PLAYER_PRIVATE

        w(ws, r, 1, f'▸ {pn}  {badge}', FONT_L1)
        for jj in range(1, nc + 1): ws.cell(row=r, column=jj).fill = pf
        c = w(ws, r, yn + 2, notes, SOURCE_FONTS.get(sq, FONT_NOTE))
        c.alignment = ALIGN_WRAP
        ws.row_dimensions[r].height = max(30, 15 * (1 + len(notes) // 60))
        r += 1

        capr = r
        caps = player.get('capacity', [])
        w(ws, r, 1, f'    Nameplate Capacity ({cfg["unit"]})', FONT_L2)
        for j in range(yn): wi(ws, r, j + 2, caps[j] if j < len(caps) else 0, FMT_COMMA, hl=(sq in ('A', 'B')))
        conf = {'A': 'Verified (10-K)', 'B': 'Reported (IR)', 'C': 'Estimated (ind. rpt)', 'D': 'Rough est.'}
        w(ws, r, yn + 2, conf.get(sq, ''), SOURCE_FONTS.get(sq, FONT_NOTE))
        r += 1

        utilr = r
        utils = player.get('utilization', [])
        w(ws, r, 1, f'    Utilization Rate', FONT_L2)
        for j in range(yn): wi(ws, r, j + 2, utils[j] if j < len(utils) else 0.0, FMT_PCT, hl=True)
        r += 1

        w(ws, r, 1, f'    Expansion Notes', FONT_NOTE)
        w(ws, r, yn + 2, 'Capex / permits / timeline', FONT_NOTE)
        r += 1

        sm[pn] = {'cap_row': capr, 'util_row': utilr}; r += 1

    # ── Price Mechanism ──
    section_row(ws, r, 'PRICE MECHANISM ASSUMPTIONS', nc); r += 1
    pm = cfg.get('demand', {}).get('price_mechanism', {})
    pm_fields = [
        ('Shortage Elasticity (ASP Δ% per 1% gap)', pm.get('shortage_elasticity', 0), 'Historical regression'),
        ('Surplus Elasticity (ASP Δ% per 1% gap)',  pm.get('surplus_elasticity', 0),  'Historical regression'),
        ('Inventory Buffer (weeks)',                  pm.get('inventory_buffer_weeks', 0), 'Channel checks'),
        ('Structural vs Cyclical [S/C]',             pm.get('structural_vs_cyclical', ''), 'Analyst judgment'),
        ('Price Floor / Marginal Cost ($)',           pm.get('price_floor_marginal_cost', 0), 'Cost curve of marginal producer'),
    ]
    for lb, val, src in pm_fields:
        w(ws, r, 1, f'  {lb}', FONT_L2)
        for j in range(yn): wi(ws, r, j + 2, val)
        w(ws, r, yn + 2, src, FONT_NOTE); r += 1

    return ws, dm, sm


# ── DEMAND ───────────────────────────────────────────────────────────────────

def build_demand(wb, cfg, dm):
    ws = wb.create_sheet('Demand')
    years = cfg['years']; yn = len(years)
    unit = cfg['unit']; ru = cfg.get('revenue_unit', '$M')
    nc = yn + 1
    set_widths(ws, [45] + [16] * yn)

    engine_mode = '_engine_vol_row' in dm
    title = 'Demand — Archetype-Driven Decomposition' if engine_mode else 'Demand — Bottom-Up Decomposition'
    r = 1; w(ws, r, 1, title, FONT_TITLE)
    subtitle = f'Vol: {unit} | Rev: {ru} | Green = Assumptions link'
    if engine_mode:
        subtitle += ' | 🔧 Engine-driven volumes'
    r = 2; w(ws, r, 1, subtitle, FONT_SUBTITLE)
    r = 4; hdr_row(ws, r, ['Metric'] + years); r += 1

    seg_rev_rows = []; seg_vol_rows = []
    ut = cfg.get('unit_type', 'stock')  # flow or stock
    af = cfg.get('annualization_factor', 1)
    rd = cfg.get('revenue_divisor', 1)

    # V3.3: Engine total volume row in Demand sheet
    if engine_mode:
        evr_assumptions = dm['_engine_vol_row']
        section_row(ws, r, '🔧 ENGINE TOTAL VOLUME', nc); r += 1
        engine_demand_row = r
        w(ws, r, 1, f'  Engine Computed Total ({unit})', FONT_L1)
        for j in range(yn):
            wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, evr_assumptions)}", FMT_COMMA, FONT_XREF)
        r += 1; yoy(ws, r, engine_demand_row, yn, '  Engine Vol YoY'); r += 2

    for seg in cfg['demand']['segments']:
        sn = seg['name']; subs = seg.get('sub_segments', []); sd = dm[sn]
        section_row(ws, r, sn, nc); r += 1

        if engine_mode:
            # V3.3: Link to allocation-computed volume from Assumptions
            avr = sd.get('_total_vol_row')
            svr = r
            w(ws, r, 1, f'  Volume ({unit}) [engine × alloc%]', FONT_L2)
            if avr:
                for j in range(yn): wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, avr)}", FMT_COMMA, FONT_XREF)
            else:
                for j in range(yn): wi(ws, r, j + 2, 0, FMT_COMMA)
            r += 1
        elif subs:
            sub_vr = []
            for sub in subs:
                sub_name = sub['name'] if isinstance(sub, dict) else sub
                sub_data = sd.get(sub_name, {})
                avr = sub_data.get('vol_row')
                vr = r; sub_vr.append(vr)
                w(ws, r, 1, f'    {sub_name} ({unit})', FONT_L3)
                if avr:
                    for j in range(yn): wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, avr)}", FMT_COMMA, FONT_XREF)
                else:
                    for j in range(yn): wi(ws, r, j + 2, 0, FMT_COMMA)
                r += 1

            svr = r
            w(ws, r, 1, f'  ↳ {sn} Total Vol ({unit})', FONT_L1)
            for j in range(yn):
                refs = '+'.join([cr(j + 2, v) for v in sub_vr])
                wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
            r += 1
        else:
            avr = sd.get('_total_vol_row')
            svr = r
            w(ws, r, 1, f'  Volume ({unit})', FONT_L2)
            if avr:
                for j in range(yn): wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, avr)}", FMT_COMMA, FONT_XREF)
            else:
                for j in range(yn): wi(ws, r, j + 2, 0, FMT_COMMA)
            r += 1

        seg_vol_rows.append(svr)
        yoy(ws, r, svr, yn, '  Vol YoY'); r += 1

        ar = r; aar = sd.get('_asp_row')
        w(ws, r, 1, f'  ASP ({cfg.get("asp_label", "$")})', FONT_L2)
        if aar:
            for j in range(yn): wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, aar)}", FMT_MONEY, FONT_XREF)
        else:
            for j in range(yn): wi(ws, r, j + 2, 0, FMT_MONEY)
        r += 1

        rr = r; seg_rev_rows.append(rr)
        w(ws, r, 1, f'  Revenue ({ru})', FONT_L1)
        # Unit Consistency: K_units × $/unit = $K; divide by 1000 to get $M
        _rev_div = 1000 if ru == '$M' and cfg.get('unit', '').startswith('K') else rd
        for j in range(yn):
            if ut == 'flow':
                wf(ws, r, j + 2, f'={cr(j+2, svr)}*{cr(j+2, ar)}*{af}/{_rev_div}', FMT_MONEY)
            else:
                wf(ws, r, j + 2, f'={cr(j+2, svr)}*{cr(j+2, ar)}/{_rev_div}', FMT_MONEY)
        r += 1

        yoy(ws, r, rr, yn, '  Rev YoY'); r += 1; r += 1

    # Total
    section_row(ws, r, 'TOTAL MARKET', nc); r += 1

    # Fix 6: Total Volume aggregation row
    tvr = r
    w(ws, r, 1, f'Total Market Volume ({unit})', FONT_L1)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, x) for x in seg_vol_rows])
        wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
    r += 1; yoy(ws, r, tvr, yn, 'Total Vol YoY', FONT_L2); r += 1

    trr = r
    w(ws, r, 1, f'Total Market Revenue ({ru})', FONT_L1)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, x) for x in seg_rev_rows])
        wf(ws, r, j + 2, f'={refs}', FMT_MONEY)
    r += 1; yoy(ws, r, trr, yn, 'Total Rev YoY', FONT_L2); r += 2

    # Cross-check
    cir = r
    w(ws, r, 1, '✓ CHECK: Top-Down Estimate (input)', FONT_CHECK)
    for j in range(yn): wi(ws, r, j + 2, 0, FMT_MONEY)
    r += 1
    w(ws, r, 1, '  Δ BU vs TD', FONT_CHECK)
    for j in range(yn):
        wf(ws, r, j + 2, f'=IF({cr(j+2, cir)}=0,0,({cr(j+2, trr)}-{cr(j+2, cir)})/{cr(j+2, cir)})', FMT_PCT, FONT_CHECK)
        ws.cell(row=r, column=j + 2).fill = FILL_CHECK

    return ws, tvr, trr, seg_rev_rows, seg_vol_rows


# ── SUPPLY ───────────────────────────────────────────────────────────────────

def build_supply(wb, cfg, sm):
    ws = wb.create_sheet('Supply')
    years = cfg['years']; yn = len(years)
    unit = cfg['unit']; nc = yn + 2
    set_widths(ws, [45] + [15] * yn + [40])

    r = 1; w(ws, r, 1, 'Supply — Capacity by Player', FONT_TITLE)
    r = 2; w(ws, r, 1, f'{unit} | Green = Assumptions link | Orange = private co.', FONT_SUBTITLE)
    r = 4

    # Player profile table
    section_row(ws, r, 'PLAYER PROFILES', nc); r += 1
    for i, h in enumerate(['Player', 'Status', 'Ticker', 'Src Grade', 'Notes']):
        w(ws, r, i + 1, h, FONT_HEADER, fill=FILL_HEADER)
        ws.cell(row=r, column=i + 1).fill = FILL_HEADER
    r += 1
    for p in cfg['supply']['players']:
        il = p.get('listed', False); pf = FILL_PLAYER_LISTED if il else FILL_PLAYER_PRIVATE
        sq = p.get('source_quality', 'C')
        w(ws, r, 1, p['name'], FONT_L2, fill=pf)
        w(ws, r, 2, 'Listed' if il else 'PRIVATE', FONT_L2, fill=pf)
        w(ws, r, 3, p.get('ticker') or '—', FONT_L2, fill=pf)
        w(ws, r, 4, f'Grade {sq}', SOURCE_FONTS.get(sq, FONT_NOTE), fill=pf)
        c = w(ws, r, 5, p.get('notes', ''), FONT_NOTE, fill=pf)
        c.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=nc)
        ws.row_dimensions[r].height = max(30, 15 * (1 + len(p.get('notes', '')) // 70))
        r += 1
    r += 1

    # Capacity detail
    hdr_row(ws, r, ['Metric'] + years + ['Notes']); r += 1
    per = []  # player effective rows (non-subset only)
    per_all = []  # all player effective rows
    player_eff_row_map = {}  # name -> row
    subset_names = {p['is_subset_of'] for p in cfg['supply']['players'] if p.get('is_subset_of')}
    subset_players = {p['name'] for p in cfg['supply']['players'] if p.get('is_subset_of')}

    for p in cfg['supply']['players']:
        pn = p['name']; il = p.get('listed', False); sq = p.get('source_quality', 'C')
        pf = FILL_PLAYER_LISTED if il else FILL_PLAYER_PRIVATE
        is_sub = p.get('is_subset_of')
        badge = f"[{p.get('ticker', 'PRIV')}][{sq}]"
        if is_sub:
            badge += f" ⊂ {is_sub}"

        w(ws, r, 1, f'▸ {pn}  {badge}', FONT_L1, fill=pf)
        for jj in range(2, nc + 1): ws.cell(row=r, column=jj).fill = pf
        r += 1

        # Capacity
        capr = r; sd = sm.get(pn, {}); acr = sd.get('cap_row')
        w(ws, r, 1, f'  Nameplate Cap ({unit})', FONT_L2)
        for j in range(yn):
            if acr: wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, acr)}", FMT_COMMA, FONT_XREF)
            else: wi(ws, r, j + 2, 0, FMT_COMMA)
        conf = {'A': 'Verified (10-K)', 'B': 'Reported (IR)', 'C': 'Est. (ind.)', 'D': 'Rough est.'}
        w(ws, r, yn + 2, conf.get(sq, ''), SOURCE_FONTS.get(sq, FONT_NOTE))
        r += 1

        # Utilization
        utilr = r; aur = sd.get('util_row')
        w(ws, r, 1, f'  Utilization', FONT_L2)
        for j in range(yn):
            if aur: wf(ws, r, j + 2, f"=Assumptions!{cr(j+2, aur)}", FMT_PCT, FONT_XREF)
            else: wi(ws, r, j + 2, 0.0, FMT_PCT)
        r += 1

        # Effective = cap × util
        effr = r
        per_all.append(effr)
        player_eff_row_map[pn] = effr
        if not is_sub:
            per.append(effr)
        w(ws, r, 1, f'  Effective Output ({unit})', FONT_L1)
        for j in range(yn): wf(ws, r, j + 2, f'={cr(j+2, capr)}*{cr(j+2, utilr)}', FMT_COMMA)
        r += 1

        # Cap addition (百分比 YoY，绝对变化用 Δ 命名)
        w(ws, r, 1, f'  Cap Δ Addition ({unit})', FONT_L3)
        w(ws, r, 2, 'n/a', FONT_NOTE)
        for j in range(1, yn): wf(ws, r, j + 2, f'={cr(j+2, capr)}-{cr(j+1, capr)}', FMT_COMMA, FONT_L3)
        r += 1
        w(ws, r, 1, f'  Cap YoY', FONT_L3)
        w(ws, r, 2, 'n/a', FONT_NOTE)
        for j in range(1, yn): wf(ws, r, j + 2, f'=IF({cr(j+1, capr)}=0,0,({cr(j+2, capr)}-{cr(j+1, capr)})/{cr(j+1, capr)})', FMT_PCT, FONT_L3)
        r += 1; r += 1

    # Totals (exclude subsets to avoid double-counting)
    section_row(ws, r, 'INDUSTRY TOTAL', nc); r += 1
    ter = r
    w(ws, r, 1, f'Total Effective Supply ({unit})', FONT_L1)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, e) for e in per])
        wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
    r += 1
    yoy(ws, r, ter, yn, 'Supply YoY', FONT_L2)

    return ws, ter, per_all, player_eff_row_map


# ── SD BALANCE ───────────────────────────────────────────────────────────────

def build_sd_balance(wb, cfg, demand_vol_row, supply_eff_row, dm=None):
    ws = wb.create_sheet('SD_Balance')
    years = cfg['years']; yn = len(years)
    unit = cfg['unit']; ru = cfg.get('revenue_unit', '$M'); nc = yn + 1
    ut = cfg.get('unit_type', 'stock')
    af = cfg.get('annualization_factor', 1)
    rd = cfg.get('revenue_divisor', 1)
    set_widths(ws, [45] + [16] * yn)

    r = 1; w(ws, r, 1, 'Supply-Demand Balance & Price Mechanism', FONT_TITLE)
    r = 2; w(ws, r, 1, 'Gap drives ASP | Green = cross-sheet link', FONT_SUBTITLE)
    r = 4; hdr_row(ws, r, ['Metric'] + years); r += 1

    section_row(ws, r, 'CORE S/D', nc); r += 1

    dr = r
    w(ws, r, 1, f'Total Demand ({unit})  ← Demand sheet', FONT_L2)
    for j in range(yn): wf(ws, r, j + 2, f"=Demand!{cr(j+2, demand_vol_row)}", FMT_COMMA, FONT_XREF)
    r += 1; yoy(ws, r, dr, yn, '  Demand YoY'); r += 1

    sr = r
    w(ws, r, 1, f'Total Supply ({unit})  ← Supply sheet', FONT_L2)
    for j in range(yn): wf(ws, r, j + 2, f"=Supply!{cr(j+2, supply_eff_row)}", FMT_COMMA, FONT_XREF)
    r += 1; yoy(ws, r, sr, yn, '  Supply YoY'); r += 1

    gr = r
    w(ws, r, 1, f'S/D Gap ({unit})  [+ surplus / - shortage]', FONT_L1)
    for j in range(yn): wf(ws, r, j + 2, f'={cr(j+2, sr)}-{cr(j+2, dr)}', FMT_COMMA)
    r += 1

    gpr = r
    w(ws, r, 1, 'Gap as % of Demand', FONT_L1)
    for j in range(yn): wf(ws, r, j + 2, f'=IF({cr(j+2, dr)}=0,0,{cr(j+2, gr)}/{cr(j+2, dr)})', FMT_PCT)
    r += 1

    w(ws, r, 1, 'Direction', FONT_L2)
    for j in range(yn):
        wf(ws, r, j + 2, f'=IF({cr(j+2, gr)}>0,"SURPLUS",IF({cr(j+2, gr)}<0,"SHORTAGE","BALANCED"))', '@')
    r += 2

    # Price
    section_row(ws, r, 'PRICE MECHANISM', nc); r += 1

    aspr = r
    asp_est = cfg.get('demand', {}).get('asp_estimates', [])
    w(ws, r, 1, f'Blended ASP ({cfg.get("asp_label", "$")})', FONT_L1)
    # Build weighted-average ASP formula from per-segment vol × asp rows in Assumptions
    # dm[sn] has '_vol_row' (total vol) and '_asp_row' for each segment
    seg_names = [s['name'] for s in cfg.get('demand', {}).get('segments', [])]
    vol_rows = [dm[sn]['_total_vol_row'] for sn in seg_names if sn in dm and '_total_vol_row' in dm[sn]]
    asp_rows = [dm[sn]['_asp_row'] for sn in seg_names if sn in dm and '_asp_row' in dm[sn]]
    for j in range(yn):
        col = get_column_letter(j + 2)
        if asp_est and j < len(asp_est):
            # Explicit override takes priority
            wi(ws, r, j + 2, asp_est[j], FMT_MONEY, hl=True)
        elif vol_rows and asp_rows and len(vol_rows) == len(asp_rows):
            # Weighted average: Σ(vol_i × asp_i) / Σ(vol_i)
            num = '+'.join(f'Assumptions!{col}{vr}*Assumptions!{col}{ar}' for vr, ar in zip(vol_rows, asp_rows))
            den = '+'.join(f'Assumptions!{col}{vr}' for vr in vol_rows)
            wf(ws, r, j + 2, f'=IF({den}=0,0,({num})/({den}))', FMT_MONEY, FONT_XREF)
        else:
            wi(ws, r, j + 2, 0, FMT_MONEY, hl=True)
    r += 1; yoy(ws, r, aspr, yn, '  ASP YoY'); r += 1

    mktr = r
    w(ws, r, 1, f'Implied Market Size ({ru})', FONT_L1)
    _sd_rev_div = 1000 if ru == '$M' and cfg.get('unit', '').startswith('K') else rd
    for j in range(yn):
        if ut == 'flow':
            wf(ws, r, j + 2, f'={cr(j+2, dr)}*{cr(j+2, aspr)}*{af}/{_sd_rev_div}', FMT_MONEY)
        else:
            wf(ws, r, j + 2, f'={cr(j+2, dr)}*{cr(j+2, aspr)}/{_sd_rev_div}', FMT_MONEY)
    r += 1; yoy(ws, r, mktr, yn, '  Mkt YoY'); r += 2

    # Regression
    section_row(ws, r, 'S/D vs PRICE SENSITIVITY', nc); r += 1
    w(ws, r, 1, 'Plot: X = Gap%, Y = ASP Δ% to find elasticity', FONT_SUBTITLE); r += 1

    w(ws, r, 1, '  Gap % (repeat)', FONT_L3)
    for j in range(yn): wf(ws, r, j + 2, f'={cr(j+2, gpr)}', FMT_PCT, FONT_L3)
    r += 1
    w(ws, r, 1, '  ASP YoY (repeat)', FONT_L3)
    w(ws, r, 2, 'n/a', FONT_NOTE)
    for j in range(1, yn):
        wf(ws, r, j + 2, f'=IF({cr(j+1, aspr)}=0,0,({cr(j+2, aspr)}-{cr(j+1, aspr)})/{cr(j+1, aspr)})', FMT_PCT, FONT_L3)
    r += 2

    pm = cfg.get('demand', {}).get('price_mechanism', {})
    w(ws, r, 1, '  Shortage Elasticity', FONT_L2); wi(ws, r, 2, pm.get('shortage_elasticity', 0), FMT_COMMA1, hl=True); r += 1
    w(ws, r, 1, '  Surplus Elasticity', FONT_L2); wi(ws, r, 2, pm.get('surplus_elasticity', 0), FMT_COMMA1, hl=True); r += 1
    w(ws, r, 1, '  Inventory Buffer (weeks)', FONT_L2); wi(ws, r, 2, pm.get('inventory_buffer_weeks', 0), FMT_COMMA); r += 1
    w(ws, r, 1, '  Price Floor ($)', FONT_L2); wi(ws, r, 2, pm.get('price_floor_marginal_cost', 0), FMT_MONEY); r += 2

    # Scenarios — from cfg; use total demand volume for market size calc
    asc = cfg.get('demand', {}).get('asp_scenarios', {})
    tvol_vals = [sum(_get_seg_vols(seg, yn)[j] for seg in cfg.get('demand', {}).get('segments', [])) for j in range(yn)]
    section_row(ws, r, 'ASP SCENARIOS', nc); r += 1
    for sc_label, sc_key in [('Base Case ASP', 'base'), ('Bull Case (tight S/D)', 'bull'), ('Bear Case (oversupply)', 'bear')]:
        w(ws, r, 1, f'  {sc_label}', FONT_L2)
        sc_vals = asc.get(sc_key, [])
        for j in range(yn): wi(ws, r, j + 2, sc_vals[j] if j < len(sc_vals) else 0, FMT_MONEY)
        r += 1

    for sc_label, sc_key in [('Base Case Mkt Size', 'base'), ('Bull Case Mkt Size', 'bull'), ('Bear Case Mkt Size', 'bear')]:
        w(ws, r, 1, f'  {sc_label} ({ru})', FONT_L2)
        sc_vals = asc.get(sc_key, [])
        for j in range(yn):
            asp_v = sc_vals[j] if j < len(sc_vals) else 0
            wi(ws, r, j + 2, round(asp_v * (tvol_vals[j] if j < len(tvol_vals) else 0), 2) if asp_v else 0, FMT_MONEY)
        r += 1

    r += 1
    w(ws, r, 1, '✓ CHECK: Implied mkt ≈ Demand total rev', FONT_CHECK)
    for j in range(yn): ws.cell(row=r, column=j + 2).fill = FILL_CHECK

    sd_refs = {'gap_pct_row': gpr, 'asp_row': aspr, 'mkt_row': mktr, 'demand_row': dr, 'supply_row': sr}
    return ws, sd_refs


# ── COMPETITION ──────────────────────────────────────────────────────────────

def build_competition(wb, cfg, player_eff_row_map):
    ws = wb.create_sheet('Competition')
    years = cfg['years']; yn = len(years)
    players = cfg['supply']['players']
    ru = cfg.get('revenue_unit', '$M'); unit = cfg['unit']; nc = yn + 1
    set_widths(ws, [45] + [16] * yn)

    r = 1; w(ws, r, 1, 'Competitive Landscape Evolution', FONT_TITLE)
    r = 2; w(ws, r, 1, f'Revenue share + Capacity share | Orange = private', FONT_SUBTITLE)
    r = 4; hdr_row(ws, r, ['Metric'] + years); r += 1

    # Revenue — P0-5: fill from revenue_estimates in cfg
    section_row(ws, r, f'REVENUE BY PLAYER ({ru})', nc); r += 1
    rvr = {}
    for p in players:
        pn = p['name']; il = p.get('listed', False)
        pf = FILL_PLAYER_LISTED if il else FILL_PLAYER_PRIVATE
        rvr[pn] = r; badge = '📊' if il else '🔒'
        rev_src = p.get('revenue_source', '')
        w(ws, r, 1, f'  {badge} {pn}', FONT_L2, fill=pf)
        rev_ests = p.get('revenue_estimates', [])
        for j in range(yn):
            val = rev_ests[j] if j < len(rev_ests) else 0
            wi(ws, r, j + 2, val or 0, FMT_MONEY)
            if not il: ws.cell(row=r, column=j + 2).fill = FILL_PLAYER_PRIVATE
        if rev_src:
            w(ws, r, yn + 2, rev_src, FONT_NOTE)
        r += 1

    trr = r
    w(ws, r, 1, f'  Total ({ru})', FONT_L1)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, rvr[p['name']]) for p in players])
        wf(ws, r, j + 2, f'={refs}', FMT_MONEY)
    r += 2

    # Share
    section_row(ws, r, 'REVENUE MARKET SHARE', nc); r += 1
    shr = {}
    for p in players:
        pn = p['name']; shr[pn] = r
        w(ws, r, 1, f'  {pn}', FONT_L2)
        for j in range(yn):
            wf(ws, r, j + 2, f'=IF({cr(j+2, trr)}=0,0,{cr(j+2, rvr[pn])}/{cr(j+2, trr)})', FMT_PCT)
        r += 1

    w(ws, r, 1, '  ✓ Sum', FONT_CHECK)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, shr[p['name']]) for p in players])
        wf(ws, r, j + 2, f'={refs}', FMT_PCT, FONT_CHECK)
        ws.cell(row=r, column=j + 2).fill = FILL_CHECK
    r += 2

    # Concentration
    section_row(ws, r, 'CONCENTRATION', nc); r += 1
    no = [p for p in players if p['name'].lower() not in ('others', 'other miners')]
    t3 = no[:3]; t5 = no[:5]

    # CR3/CR5：用连续 range 的 LARGE()，避免数组常量（Excel iOS 不支持）
    shr_rows = [shr[p['name']] for p in no if p['name'] in shr]
    if len(shr_rows) >= 3:
        first_shr, last_shr = shr_rows[0], shr_rows[-1]
        w(ws, r, 1, '  CR3 (top-3 by revenue share, dynamic)', FONT_L2)
        for j in range(yn):
            col_letter = col(j + 2)
            rng = f'{col_letter}{first_shr}:{col_letter}{last_shr}'
            wf(ws, r, j + 2, f'=LARGE({rng},1)+LARGE({rng},2)+LARGE({rng},3)', FMT_PCT)
        r += 1
    if len(shr_rows) >= 5:
        w(ws, r, 1, '  CR5 (top-5 by revenue share, dynamic)', FONT_L2)
        for j in range(yn):
            col_letter = col(j + 2)
            rng = f'{col_letter}{first_shr}:{col_letter}{last_shr}'
            wf(ws, r, j + 2, f'=LARGE({rng},1)+LARGE({rng},2)+LARGE({rng},3)+LARGE({rng},4)+LARGE({rng},5)', FMT_PCT)
        r += 1

    w(ws, r, 1, '  HHI (x10000)', FONT_L2)
    for j in range(yn):
        refs = '+'.join([f'({cr(j+2, shr[p["name"]])})^2' for p in players])
        wf(ws, r, j + 2, f'=({refs})*10000', FMT_COMMA)
    r += 2

    # Share change
    section_row(ws, r, 'SHARE CHANGE (YoY ppts)', nc); r += 1
    for p in players:
        pn = p['name']
        w(ws, r, 1, f'  {pn} Δ', FONT_L3)
        w(ws, r, 2, 'n/a', FONT_NOTE)
        for j in range(1, yn):
            wf(ws, r, j + 2, f'={cr(j+2, shr[pn])}-{cr(j+1, shr[pn])}', FMT_PCT, FONT_L3)
        r += 1
    r += 1

    # Capacity share — cross-sheet from Supply
    section_row(ws, r, f'CAPACITY SHARE (Eff. Output, {unit})', nc); r += 1
    copr = {}
    for p in players:
        pn = p['name']; copr[pn] = r
        sup_row = player_eff_row_map.get(pn)
        w(ws, r, 1, f'  {pn} Output', FONT_L2)
        if sup_row:
            for j in range(yn): wf(ws, r, j + 2, f"=Supply!{cr(j+2, sup_row)}", FMT_COMMA, FONT_XREF)
        else:
            for j in range(yn): wi(ws, r, j + 2, 0, FMT_COMMA, FONT_XREF)
        r += 1

    tcr = r
    w(ws, r, 1, '  Total Output', FONT_L1)
    for j in range(yn):
        refs = '+'.join([cr(j + 2, copr[p['name']]) for p in players])
        wf(ws, r, j + 2, f'={refs}', FMT_COMMA)
    r += 1

    cap_shr = {}
    for p in players:
        pn = p['name']; cap_shr[pn] = r
        w(ws, r, 1, f'  {pn} Cap Share', FONT_L3)
        for j in range(yn):
            wf(ws, r, j + 2, f'=IF({cr(j+2, tcr)}=0,0,{cr(j+2, copr[pn])}/{cr(j+2, tcr)})', FMT_PCT)
        r += 1
    r += 1

    # Barriers — P2-13: fill from cfg.competitive_barriers
    section_row(ws, r, 'COMPETITIVE BARRIERS (1-5)', nc); r += 1
    cb = cfg.get('competitive_barriers', {})
    barrier_map = [
        ('Technology / Know-How', cb.get('technology', 0)),
        ('Scale / Cost Curve', cb.get('scale_cost', 0)),
        ('Customer Lock-In', cb.get('customer_lock_in', 0)),
        ('Capital Intensity', cb.get('capital_intensity', 0)),
        ('Regulatory / License', cb.get('regulatory', 0)),
        ('Resource Access', cb.get('resource_access', 0)),
    ]
    bs = r
    for b_label, b_val in barrier_map:
        w(ws, r, 1, f'  {b_label}', FONT_L2); wi(ws, r, 2, b_val or 0, FMT_COMMA); r += 1
    w(ws, r, 1, '  TOTAL (/30)', FONT_L1)
    wf(ws, r, 2, f'=SUM(B{bs}:B{r-1})', FMT_COMMA); r += 2

    # New entrant tracker
    section_row(ws, r, 'NEW ENTRANT / EXPANSION TRACKER', nc); r += 1
    for i, h in enumerate(['Player', 'New Capacity', 'Online Date', 'Credibility (H/M/L)', 'Status / Notes']):
        w(ws, r, i + 1, h, FONT_HEADER, fill=FILL_HEADER)
        ws.cell(row=r, column=i + 1).fill = FILL_HEADER
    r += 1
    for _ in range(8):
        for cc in range(1, 6): wi(ws, r, cc, '', None, FONT_INPUT)
        r += 1

    # 隐藏辅助行：玩家名称（供 Summary Leader 公式用，避免数组常量）
    name_helper_row = r
    w(ws, r, 1, '_player_names_helper', FONT_NOTE)
    ws.row_dimensions[r].hidden = True
    for i, p in enumerate(players):
        ws.cell(row=r, column=i + 2, value=p['name'])
    r += 1

    comp_refs = {'rev_share': shr, 'cap_share': cap_shr, 'hhi_row': None,
                 'total_rev_row': trr, 'name_helper_row': name_helper_row}

    # ── Conditional Formatting: market share block ────────────────────────────
    share_start = min(shr.values())
    share_end   = max(shr.values())
    col_end     = get_column_letter(yn + 1)
    share_range = f"B{share_start}:{col_end}{share_end}"

    ws.conditional_formatting.add(
        share_range,
        ColorScaleRule(
            start_type="min",        start_color="63BE7B",
            mid_type="percentile",   mid_value=50, mid_color="FFEB84",
            end_type="max",          end_color="F8696B"
        )
    )

    # Bold border around share block (include header row above and sum row below)
    thick = Side(style="medium"); thin = Side(style="thin")
    for row_idx in range(share_start - 1, share_end + 2):
        for col_idx in range(1, yn + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(
                top    = thick if row_idx == share_start - 1 else thin,
                bottom = thick if row_idx == share_end + 1   else thin,
                left   = thick if col_idx == 1               else thin,
                right  = thick if col_idx == yn + 1          else thin,
            )

    return ws, comp_refs


# ── SUMMARY ──────────────────────────────────────────────────────────────────

def build_summary(wb, cfg, sd_refs, comp_refs):
    ws = wb.create_sheet('Summary', 0)
    years = cfg['years']; yn = len(years)
    ru = cfg.get('revenue_unit', '$M'); nc = yn + 1
    set_widths(ws, [42] + [16] * yn)

    r = 1; w(ws, r, 1, cfg['title'], FONT_TITLE)
    r = 2; w(ws, r, 1, 'Dashboard | Captain Market Sizing Skill v2.1', FONT_SUBTITLE)
    r = 4; hdr_row(ws, r, ['Dashboard'] + years); r += 1

    # Market Size — cross-sheet from SD_Balance
    section_row(ws, r, 'MARKET SIZE', nc); r += 1
    mkt_row = sd_refs.get('mkt_row')
    asp_row = sd_refs.get('asp_row')

    w(ws, r, 1, f'Total Market Revenue ({ru})', FONT_L2)
    if mkt_row:
        for j in range(yn): wf(ws, r, j + 2, f"=SD_Balance!{cr(j+2, mkt_row)}", FMT_MONEY, FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, 0, FMT_MONEY, FONT_XREF)
    rev_r = r; r += 1

    w(ws, r, 1, 'Revenue YoY', FONT_L2)
    w(ws, r, 2, 'n/a', FONT_NOTE)
    for j in range(1, yn):
        wf(ws, r, j + 2, f'=IF({cr(j+1, rev_r)}=0,0,({cr(j+2, rev_r)}-{cr(j+1, rev_r)})/{cr(j+1, rev_r)})', FMT_PCT, FONT_XREF)
    r += 1

    w(ws, r, 1, f'Blended ASP ({cfg.get("asp_label", "$")})', FONT_L2)
    if asp_row:
        for j in range(yn): wf(ws, r, j + 2, f"=SD_Balance!{cr(j+2, asp_row)}", FMT_MONEY, FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, 0, FMT_MONEY, FONT_XREF)
    asp_r = r; r += 1

    w(ws, r, 1, 'ASP YoY', FONT_L2)
    w(ws, r, 2, 'n/a', FONT_NOTE)
    for j in range(1, yn):
        wf(ws, r, j + 2, f'=IF({cr(j+1, asp_r)}=0,0,({cr(j+2, asp_r)}-{cr(j+1, asp_r)})/{cr(j+1, asp_r)})', FMT_PCT, FONT_XREF)
    r += 2

    # Supply-Demand — cross-sheet from SD_Balance
    section_row(ws, r, 'SUPPLY-DEMAND', nc); r += 1
    gp_row = sd_refs.get('gap_pct_row')

    w(ws, r, 1, 'S/D Gap (%)', FONT_L2)
    if gp_row:
        for j in range(yn): wf(ws, r, j + 2, f"=SD_Balance!{cr(j+2, gp_row)}", FMT_PCT, FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, 0, FMT_PCT, FONT_XREF)
    r += 1

    w(ws, r, 1, 'Direction', FONT_L2)
    if gp_row:
        for j in range(yn): wf(ws, r, j + 2, f'=IF(SD_Balance!{cr(j+2, gp_row)}>0,"SURPLUS",IF(SD_Balance!{cr(j+2, gp_row)}<0,"SHORTAGE","BALANCED"))', '@', FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, '', '@', FONT_XREF)
    r += 2

    # Competition — cross-sheet from Competition
    section_row(ws, r, 'COMPETITION', nc); r += 1
    rev_share = comp_refs.get('rev_share', {})
    players = cfg['supply']['players']
    no = [p for p in players if p['name'].lower() not in ('others', 'other miners')]
    t3 = no[:3]

    w(ws, r, 1, 'CR3', FONT_L2)
    if t3 and rev_share:
        for j in range(yn):
            refs = '+'.join([f"Competition!{cr(j+2, rev_share[p['name']])}" for p in t3 if p['name'] in rev_share])
            wf(ws, r, j + 2, f'={refs}', FMT_PCT, FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, 0, FMT_PCT, FONT_XREF)
    r += 1

    w(ws, r, 1, 'Leader', FONT_L2)
    name_helper_row = comp_refs.get('name_helper_row')
    shr_rows_list = [rev_share[p['name']] for p in players if p['name'] in rev_share]
    if name_helper_row and shr_rows_list and rev_share:
        first_shr_s = shr_rows_list[0]; last_shr_s = shr_rows_list[-1]
        n_players = len(players)
        for j in range(yn):
            col_letter = col(j + 2)
            name_rng = f'Competition!B{name_helper_row}:{col(n_players + 1)}{name_helper_row}'
            share_rng = f'Competition!{col_letter}{first_shr_s}:Competition!{col_letter}{last_shr_s}'
            # INDEX(name_range_row, 1, MATCH(MAX(share_col), share_col, 0))
            wf(ws, r, j + 2,
               f'=INDEX(Competition!B{name_helper_row}:{col(n_players+1)}{name_helper_row},1,MATCH(MAX({share_rng}),{share_rng},0))',
               '@', FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, players[0]['name'] if players else '', '@', FONT_XREF)
    r += 1

    w(ws, r, 1, 'Leader Share', FONT_L2)
    leader = players[0]['name'] if players else None
    if leader and leader in rev_share:
        for j in range(yn): wf(ws, r, j + 2, f"=Competition!{cr(j+2, rev_share[leader])}", FMT_PCT, FONT_XREF)
    else:
        for j in range(yn): wi(ws, r, j + 2, 0, FMT_PCT, FONT_XREF)
    r += 2

    section_row(ws, r, 'INVESTMENT CONCLUSION', nc); r += 1
    ic = cfg.get('investment_conclusion', {})
    ic_fields = [
        ('Industry Attractiveness (1-5)', str(ic.get('attractiveness', ''))),
        ('Best Window', ic.get('best_window', '')),
        ('Upside Catalyst', ic.get('upside_catalyst', '')),
        ('Downside Risk', ic.get('downside_risk', '')),
        ('Proxy Tickers', ic.get('proxy_tickers', '')),
        ('Conviction (A/B/C/D)', ic.get('conviction', '')),
        ('Notes', ic.get('notes', '')),
    ]
    for label, val in ic_fields:
        w(ws, r, 1, label, FONT_L1)
        c = ws.cell(row=r, column=2, value=val or '')
        c.font = FONT_INPUT; c.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=yn + 1)
        r += 1

    r += 1
    w(ws, r, 1, '✅ All green values auto-linked to SD_Balance/Competition', FONT_CHECK)
    return ws


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description='Market Sizing Model v3.2')
    p.add_argument('--config', required=True, help='JSON config file')
    p.add_argument('--output', default=None, help='Output xlsx path')
    p.add_argument('--strict', action='store_true', help='Abort on any L3+ issue')
    args = p.parse_args()

    with open(args.config) as f:
        cfg = json.load(f)

    # ── V3.2 Pre-flight Gate ─────────────────────────────────────────────────
    from validity_engine import preflight_check, validate_all, issues_to_legacy, Severity
    pf_issues = preflight_check(cfg)
    l4_blockers = [i for i in pf_issues if i.severity == Severity.L4_VALIDITY]
    if l4_blockers:
        print('🛑 PRE-FLIGHT FAILED — model generation blocked:')
        for i in l4_blockers:
            print(f'   L4 [{i.rule}] {i.message}')
        print(f'\n   Fix {len(l4_blockers)} L4 issues in config before generating.')
        sys.exit(1)

    slug = cfg['title'].replace(' ', '_')[:30]
    output = args.output or f'/mnt/user-data/outputs/market_sizing_{slug}.xlsx'

    wb = Workbook(); wb.remove(wb.active)

    ws_a, dm, sm = build_assumptions(wb, cfg)
    ws_d, demand_total_vol_row, demand_total_rev_row, seg_rev_rows, seg_vol_rows = build_demand(wb, cfg, dm)
    ws_s, supply_total_eff_row, player_eff_rows, player_eff_row_map = build_supply(wb, cfg, sm)
    ws_sd, sd_refs = build_sd_balance(wb, cfg, demand_total_vol_row, supply_total_eff_row, dm=dm)
    ws_c, comp_refs = build_competition(wb, cfg, player_eff_row_map)
    ws_sum = build_summary(wb, cfg, sd_refs, comp_refs)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B5'

    # Force Excel to recalculate all formulas on open
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    build_data_sheet(wb, cfg,
                     demand_total_rev_row=demand_total_rev_row,
                     supply_total_eff_row=supply_total_eff_row,
                     seg_rev_rows=seg_rev_rows,
                     seg_vol_rows=seg_vol_rows,
                     player_eff_row_map=player_eff_row_map)

    # ── V3.2: Run both legacy + engine validation ────────────────────────────
    legacy_fails, legacy_warns = validate_model(wb, cfg, sd_refs, comp_refs)
    v32_issues, validity = validate_all(cfg)
    v32_fails, v32_warns = issues_to_legacy(v32_issues)

    # Merge (deduplicate by message substring)
    all_fails = list(dict.fromkeys(legacy_fails + v32_fails))
    all_warns = list(dict.fromkeys(legacy_warns + v32_warns))

    # Build structured Audit sheet (replaces WARNINGS)
    build_audit_sheet(wb, v32_issues, validity, cfg)

    # --strict mode: abort on L3+
    if args.strict:
        l3_plus = [i for i in v32_issues if i.severity >= Severity.L3_MECHANISM]
        if l3_plus:
            print(f'🛑 STRICT MODE: {len(l3_plus)} L3+ issues — not saving.')
            for i in l3_plus:
                print(f'   L{i.severity.value} [{i.rule}] {i.message}')
            sys.exit(1)

    wb.save(output)

    total_subs = sum(len(s.get('sub_segments', [])) for s in cfg['demand']['segments'])
    listed = sum(1 for p in cfg['supply']['players'] if p.get('listed'))
    private = len(cfg['supply']['players']) - listed

    print(f'✅ Saved: {output}')
    print(f'   {cfg["title"]}')
    print(f'   Sheets: {", ".join(wb.sheetnames)}')
    print(f'   Years: {len(cfg["years"])} ({cfg["years"][0]}–{cfg["years"][-1]})')
    print(f'   Demand: {len(cfg["demand"]["segments"])} segments, {total_subs} sub-segments')
    print(f'   Supply: {len(cfg["supply"]["players"])} players ({listed} listed, {private} private)')
    print(f'   Validity: structural={validity["structural"]} | mechanical={validity["mechanical"]} | economic={validity["economic"]}')
    if all_fails:
        print(f'\n🔴 FAIL ({len(all_fails)}):')
        for m in all_fails: print(f'   • {m}')
    if all_warns:
        print(f'\n🟡 WARN ({len(all_warns)}):')
        for m in all_warns: print(f'   • {m}')
    if not all_fails and not all_warns:
        print('\n✅ All validation checks passed')


def build_audit_sheet(wb, issues, validity, cfg):
    """V3.2 Audit sheet — structured replacement for WARNINGS."""
    from validity_engine import Severity
    if 'WARNINGS' in wb.sheetnames:
        del wb['WARNINGS']
    if 'Audit' in wb.sheetnames:
        del wb['Audit']
    ws = wb.create_sheet('Audit')
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 90

    SEVERITY_COLORS = {
        Severity.L4_VALIDITY: 'CC0000',
        Severity.L3_MECHANISM: 'CC6600',
        Severity.L2_COMMENTARY: '999900',
        Severity.L1_COSMETIC: '888888',
    }

    r = 1
    ws.cell(row=r, column=1, value='MODEL AUDIT — V3.2 Validity Engine').font = Font(name='Arial', size=14, bold=True, color='1F3864')
    r = 2
    ws.cell(row=r, column=1, value=cfg.get('title', '')).font = Font(name='Arial', size=10, color='666666', italic=True)

    # Validity summary
    r = 4
    for k, v in validity.items():
        color = '007700' if v in ('valid', 'high') else 'CC6600' if v == 'medium' else 'CC0000'
        ws.cell(row=r, column=1, value=k.capitalize()).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=r, column=2, value=v.upper()).font = Font(name='Arial', size=10, bold=True, color=color)
        r += 1

    # Override log
    r += 1
    ws.cell(row=r, column=1, value='OVERRIDE LOG').font = Font(name='Arial', size=11, bold=True, color='1F3864')
    r += 1
    for o in cfg.get('field_overrides', []):
        ws.cell(row=r, column=1, value=o.get('type', '')).font = Font(name='Arial', size=9)
        ws.cell(row=r, column=2, value=o.get('scope', '')).font = Font(name='Arial', size=9)
        ws.cell(row=r, column=3, value=o.get('reason', '')).font = Font(name='Arial', size=9, color='666666')
        r += 1
    if not cfg.get('field_overrides'):
        ws.cell(row=r, column=1, value='(none declared)').font = Font(name='Arial', size=9, color='888888')
        r += 1

    # Issues by severity
    r += 1
    ws.cell(row=r, column=1, value='ALL ISSUES').font = Font(name='Arial', size=11, bold=True, color='1F3864')
    r += 1
    for hdr, col_idx in [('Severity', 1), ('Rule', 2), ('Message', 3)]:
        c = ws.cell(row=r, column=col_idx, value=hdr)
        c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3864')
    r += 1

    sorted_issues = sorted(issues, key=lambda i: -i.severity)
    for i in sorted_issues:
        color = SEVERITY_COLORS.get(i.severity, '000000')
        ws.cell(row=r, column=1, value=f'L{i.severity.value}').font = Font(name='Arial', size=9, bold=True, color=color)
        ws.cell(row=r, column=2, value=i.rule).font = Font(name='Arial', size=9, color=color)
        ws.cell(row=r, column=3, value=i.message).font = Font(name='Arial', size=9, color=color)
        r += 1

    if not issues:
        ws.cell(row=r, column=1, value='All checks passed').font = Font(name='Arial', size=12, bold=True, color='007700')

    wb.move_sheet('Audit', offset=-len(wb.sheetnames) + 1)
    return ws


def build_data_sheet(wb, cfg, demand_total_rev_row=None, supply_total_eff_row=None,
                     seg_rev_rows=None, seg_vol_rows=None, player_eff_row_map=None):
    """
    P1-10: Data sheet 用公式引用 Demand/Supply sheet，保持联动。
    """
    ws = wb.create_sheet('Data')
    years = cfg['years']; yn = len(years)
    ws.cell(row=1, column=1, value='Key Metrics — formula-linked to Demand/Supply sheets')
    ws.cell(row=2, column=1, value='year')
    for j, yr in enumerate(years):
        ws.cell(row=2, column=j + 2, value=yr)

    r = 3
    segments = cfg['demand']['segments']

    # Per-segment revenue and volume (formula refs to Demand sheet)
    for i, seg in enumerate(segments):
        sn = seg['name']
        ws.cell(row=r, column=1, value=f"demand.{sn}.volume")
        ws.cell(row=r + 1, column=1, value=f"demand.{sn}.revenue_M")
        for j in range(yn):
            if seg_vol_rows and i < len(seg_vol_rows):
                ws.cell(row=r, column=j + 2, value=f"=Demand!{cr(j+2, seg_vol_rows[i])}")
            if seg_rev_rows and i < len(seg_rev_rows):
                ws.cell(row=r + 1, column=j + 2, value=f"=Demand!{cr(j+2, seg_rev_rows[i])}")
        r += 2

    # Total demand revenue
    ws.cell(row=r, column=1, value='demand.total_revenue_M')
    for j in range(yn):
        if demand_total_rev_row:
            ws.cell(row=r, column=j + 2, value=f"=Demand!{cr(j+2, demand_total_rev_row)}")
    r += 1

    # Supply: per-player effective capacity
    players = cfg['supply']['players']
    for player in players:
        pn = player['name']
        ws.cell(row=r, column=1, value=f"supply.{pn}.effective")
        for j in range(yn):
            if player_eff_row_map and pn in player_eff_row_map:
                ws.cell(row=r, column=j + 2, value=f"=Supply!{cr(j+2, player_eff_row_map[pn])}")
        r += 1

    # Total supply
    ws.cell(row=r, column=1, value='supply.total_effective')
    for j in range(yn):
        if supply_total_eff_row:
            ws.cell(row=r, column=j + 2, value=f"=Supply!{cr(j+2, supply_total_eff_row)}")
    r += 1

    return ws

    return ws


def validate_model(wb, cfg, sd_refs, comp_refs):
    """
    Post-generation validation gate (RULE 1-12, v2).
    Returns (fail_list, warn_list). Writes WARNINGS sheet.
    """
    fails = []
    warns = []
    yn = len(cfg['years'])
    segments = cfg.get('demand', {}).get('segments', [])
    players = cfg.get('supply', {}).get('players', [])
    mid = yn // 2

    # ── Shared computed values ────────────────────────────────────────────────
    total_demand = [sum(_get_seg_vols(seg, yn)[j] for seg in segments) for j in range(yn)]
    total_supply = [
        sum((p.get('capacity', [])[j] if j < len(p.get('capacity', [])) else 0) *
            (p.get('utilization', [])[j] if j < len(p.get('utilization', [])) else 0)
            for p in players)
        for j in range(yn)
    ]
    total_vol_mid = total_demand[mid]
    blended_asp = 0
    if total_vol_mid > 0:
        blended_asp = sum(
            _get_seg_vols(seg, yn)[mid] *
            (seg.get('asp_estimates', cfg.get('demand', {}).get('asp_estimates', []))[mid]
             if mid < len(seg.get('asp_estimates', cfg.get('demand', {}).get('asp_estimates', []))) else 0)
            for seg in segments
        ) / total_vol_mid
    # revenue_unit scale: K units × $/unit = $K; if revenue_unit is $M, divide by 1000
    _ru = cfg.get('revenue_unit', '$M')
    _rev_scale = 1000 if _ru == '$M' else 1  # $K → $M
    demand_rev_mid = total_vol_mid * blended_asp / _rev_scale
    comp_rev_mid = sum(
        (p.get('revenue_estimates', [])[mid] if mid < len(p.get('revenue_estimates', [])) else 0)
        for p in players
    )

    # ── RULE 1 (v2): Volume + ASP + Revenue 各自独立锚点校验 ─────────────────
    anchors = cfg.get('demand', {}).get('anchors', {})
    # Volume anchor: compare top-1 player model volume vs real
    top1_real_vol = anchors.get('top1_player_volume')
    if top1_real_vol and players:
        top1_caps = players[0].get('capacity', [])
        top1_model_vol = top1_caps[mid] if mid < len(top1_caps) else 0
        if top1_model_vol > 0:
            ratio = max(top1_model_vol, top1_real_vol) / min(top1_model_vol, top1_real_vol)
            if ratio > 5:
                fails.append(f'[RULE 1] VOLUME ANCHOR FAIL: top-1 player model={top1_model_vol:.0f} vs real={top1_real_vol:.0f} (ratio={ratio:.1f}x > 5x)')
    # ASP anchor
    real_asp = anchors.get('typical_asp')
    if real_asp and blended_asp > 0:
        ratio = max(blended_asp, real_asp) / min(blended_asp, real_asp)
        if ratio > 5:
            fails.append(f'[RULE 1] ASP ANCHOR FAIL: model ASP={blended_asp:.2f} vs real={real_asp:.2f} (ratio={ratio:.1f}x > 5x)')
    # Revenue anchor (also checked in RULE 5, but flag here too)
    real_rev_m = anchors.get('industry_revenue_m', 0)
    if real_rev_m > 0 and demand_rev_mid > 0:
        ratio = max(demand_rev_mid, real_rev_m) / min(demand_rev_mid, real_rev_m)
        if ratio > 5:
            fails.append(f'[RULE 1] REVENUE ANCHOR FAIL: model={demand_rev_mid:.0f}$M vs anchor={real_rev_m:.0f}$M (ratio={ratio:.1f}x > 5x)')

    # ── RULE 2 (v2): S/D Ratio ───────────────────────────────────────────────
    sd_fail_years = [
        f"{cfg['years'][j]}(ratio={total_supply[j]/total_demand[j]:.2f})"
        for j in range(yn)
        if total_demand[j] > 0 and total_supply[j] > 0 and
           (total_supply[j] / total_demand[j] > 1.5 or total_supply[j] / total_demand[j] < 0.5)
    ]
    if sd_fail_years:
        fails.append(f'[RULE 2] S/D RATIO FAIL: {", ".join(sd_fail_years[:3])} — 偏离 0.5-1.5 范围')

    # ── RULE 3 (v2): Demand Revenue vs Competition Revenue ───────────────────
    if demand_rev_mid > 0 and comp_rev_mid > 0:
        delta = abs(comp_rev_mid - demand_rev_mid) / demand_rev_mid
        if delta > 0.20:
            fails.append(f'[RULE 3] REVENUE DELTA FAIL: Demand={demand_rev_mid:.0f} vs Competition={comp_rev_mid:.0f} ({delta:.0%} > 20%)')

    # ── RULE 4 (v2): Segment coverage ≥4 + Supply-Demand correspondence ──────
    if len(segments) < 4:
        fails.append(f'[RULE 4] SEGMENT COUNT FAIL: 只有 {len(segments)} 个 segment（要求 ≥4）')

    # ── RULE 4b: Supply player count ≥6 for meaningful competition analysis ──────
    players = cfg.get('supply', {}).get('players', [])
    if len(players) < 6:
        warns.append(f'[RULE 4b] SUPPLY PLAYER COUNT: 只有 {len(players)} 个供应商（建议 ≥6 才能做有效竞争分析）')

    # ── RULE 5 (v2): Top-Down 强制非零 ───────────────────────────────────────
    td_estimates = cfg.get('demand', {}).get('top_down_estimates', [])
    if not td_estimates or all(e.get('value', 0) == 0 for e in td_estimates):
        fails.append('[RULE 5] TOP-DOWN MISSING: top_down_estimates 全为空/0（强制项，不允许留 0）')
    else:
        mid_year = cfg['years'][mid]
        for td in td_estimates:
            # Only compare TD to BU for the same year (or if no year specified, use all)
            td_year = td.get('year')
            if td_year and str(td_year) != str(mid_year):
                continue
            td_val_m = td.get('value', 0) * (1000 if td.get('unit', '$B') == '$B' else 1)
            if td_val_m > 0 and demand_rev_mid > 0:
                ratio = demand_rev_mid / td_val_m
                if ratio > 2 or ratio < 0.5:
                    fails.append(f'[RULE 5] BU vs TD FAIL: BU={demand_rev_mid:.0f}$M vs TD={td_val_m:.0f}$M ({mid_year}, Δ={abs(ratio-1):.0%} > 100%)')
                elif ratio > 1.5 or ratio < 0.67:
                    warns.append(f'[RULE 5] BU vs TD WARN: BU={demand_rev_mid:.0f}$M vs TD={td_val_m:.0f}$M ({mid_year}, Δ={abs(ratio-1):.0%} > 50%)')

    # ── RULE 6 (v2): Sub-item depth ≥2 ──────────────────────────────────────
    for seg in segments:
        if len(seg.get('sub_segments', [])) < 2:
            fails.append(f'[RULE 6] SUB-ITEM DEPTH FAIL: "{seg["name"]}" 只有 {len(seg.get("sub_segments",[]))} 个 sub-item')

    # ── RULE 7 (v2): ASP differentiation + no flat trend ─────────────────────
    all_asps_mid = []
    for seg in segments:
        asps = seg.get('asp_estimates', cfg.get('demand', {}).get('asp_estimates', []))
        if mid < len(asps):
            all_asps_mid.append(asps[mid])
        # Check flat trend
        if len(asps) >= 4:
            forecast_asps = asps[mid:]
            if len(set(round(v, 4) for v in forecast_asps)) == 1:
                warns.append(f'[RULE 7] ASP FLAT WARN: "{seg["name"]}" forecast ASP 全年不变 ({forecast_asps[0]})')
    if len(all_asps_mid) >= 2 and len(set(round(v, 2) for v in all_asps_mid)) == 1:
        fails.append(f'[RULE 7] ASP DIFFERENTIATION FAIL: 所有 segment ASP 完全相同 ({all_asps_mid[0]})')

    # ── RULE 8 (v2): Entity dedup ────────────────────────────────────────────
    tickers_seen = {}
    for p in players:
        t = (p.get('ticker') or '').strip().upper()
        if t and t not in ('NONE', '—', 'NULL', ''):
            if t in tickers_seen:
                fails.append(f'[RULE 8] ENTITY DEDUP FAIL: ticker {t} 重复 ("{p["name"]}" vs "{tickers_seen[t]}")')
            else:
                tickers_seen[t] = p['name']

    # ── RULE 9 (v2): Company metadata ────────────────────────────────────────
    for p in players:
        listed = p.get('listed', False)
        ticker = (p.get('ticker') or '').strip()
        if listed and ticker in ('', 'None', '—', 'null'):
            warns.append(f'[RULE 9] METADATA WARN: "{p["name"]}" Listed 但 ticker 为空')
        if not listed and ticker and ticker not in ('None', '—', 'null', ''):
            warns.append(f'[RULE 9] METADATA WARN: "{p["name"]}" PRIVATE 但有 ticker {ticker}')

    # ── RULE 10 (v2): Utilization differentiation ────────────────────────────
    if len(players) >= 3:
        util_trends = [
            ('up' if p.get('utilization', [])[-1] > p.get('utilization', [])[0] else
             'down' if p.get('utilization', [])[-1] < p.get('utilization', [])[0] else 'flat')
            for p in players if len(p.get('utilization', [])) >= 2
        ]
        if len(set(util_trends)) == 1 and len(util_trends) >= 3:
            fails.append(f'[RULE 10] UTILIZATION FAIL: 所有 {len(util_trends)} 个 player 走势完全相同 ({util_trends[0]})')

    # ── RULE 11 (v2): Placeholder zeros ──────────────────────────────────────
    ws_a = wb['Assumptions'] if 'Assumptions' in wb.sheetnames else None
    if ws_a:
        pm_zero = [str(row[0].value).strip() for row in ws_a.iter_rows()
                   if any(kw in str(row[0].value or '') for kw in ('Elasticity', 'Inventory', 'Structural', 'Price Floor'))
                   and all(c.value in (None, 0, '') for c in row[1:yn + 1])]
        if pm_zero:
            warns.append(f'[RULE 11] PLACEHOLDER: Price Mechanism 未填充 ({len(pm_zero)} 行)')
    ws_sd = wb['SD_Balance'] if 'SD_Balance' in wb.sheetnames else None
    if ws_sd:
        sc_zero = [str(row[0].value).strip() for row in ws_sd.iter_rows()
                   if any(kw in str(row[0].value or '') for kw in ('Scenario', 'Bull', 'Bear'))
                   and all(c.value in (None, 0, '') for c in row[1:yn + 1])]
        if sc_zero:
            warns.append(f'[RULE 11] PLACEHOLDER: ASP Scenarios 未填充 ({len(sc_zero)} 行)')

    # ── v3.0 Gate 0: Objective Classifier completeness ──────────────────────
    oc = cfg.get('objective_classifier', {})
    arch = cfg.get('primary_archetype', '')
    fc = cfg.get('formula_contract', {})
    if not oc.get('sizing_objective') and not cfg.get('sizing_objective'):
        fails.append('[GATE 0] OBJECTIVE MISSING: sizing_objective 未填写')
    if not arch:
        fails.append('[GATE 0] ARCHETYPE MISSING: primary_archetype 未填写')
    if not fc.get('volume_driver'):
        fails.append('[GATE 0] FORMULA CONTRACT MISSING: formula_contract.volume_driver 未填写')
    if not cfg.get('measurement_basis'):
        fails.append('[GATE 0] MEASUREMENT BASIS MISSING: measurement_basis 未填写 (realized_reported / normalized_run_rate / demand_pool)')
    if not cfg.get('billing_unit'):
        warns.append('[GATE 0] BILLING UNIT MISSING: billing_unit 未填写 (seat/device/shipment/ton/kWh/etc.)')
    if not fc.get('price_driver'):
        warns.append('[GATE 0] ASP MECHANISM MISSING: formula_contract.price_driver 未填写')
    if not cfg.get('archetype_rationale'):
        warns.append('[GATE 0] ARCHETYPE RATIONALE MISSING: 未说明为何选此 archetype')

    # ── v3.1 Gate 0B: Archetype Dominance ────────────────────────────────────
    ad = cfg.get('archetype_decomposition', {})
    if arch and ad:
        drivers = [ad.get('volume_driver',''), ad.get('price_driver',''), ad.get('timing_driver','')]
        primary_count = sum(1 for d in drivers if d and arch.lower().replace('_','') in d.lower().replace('_',''))
        if primary_count < 2:
            warns.append(f'[GATE 0B] ARCHETYPE DOMINANCE: primary_archetype={arch} 未主导 volume/price/timing 中至少两项 (当前主导 {primary_count}/3)')

    # ── v3.1 Gate 0C: Minimum State Variables ────────────────────────────────
    msv = cfg.get('minimum_state_variables', {})
    required_by_archetype = {
        'population': ['target_base', 'penetration', 'units_per_adopter'],
        'installed_base': ['compatible_installed_base', 'attach_rate', 'replacement_rate', 'eligibility_ceiling'],
        'throughput': ['node_base', 'utilization', 'throughput_per_node', 'price_per_unit'],
        'project': ['project_pipeline', 'conversion', 'avg_ticket', 'recognition_timing'],
        'substitution': ['legacy_base', 'substitution_rate', 'switching_friction', 'relative_asp_bridge'],
        'commodity': ['demand_volume', 'supply_volume', 'inventory_buffer', 'price_formation_rule'],
    }
    if arch and arch in required_by_archetype:
        missing_vars = [v for v in required_by_archetype[arch] if not msv.get(v)]
        if missing_vars:
            fails.append(f'[GATE 0C] MINIMUM STATE VARIABLES MISSING for {arch}: {missing_vars}')

    # ── v3.1 Gate 8: Commodity Basis Consistency ─────────────────────────────
    if arch == 'commodity':
        cb = cfg.get('commodity_basis', {})
        if not cb.get('demand_volume_basis'):
            fails.append('[GATE 8] COMMODITY BASIS: demand_volume_basis 未填写')
        if not cb.get('supply_volume_basis'):
            fails.append('[GATE 8] COMMODITY BASIS: supply_volume_basis 未填写')
        if not cb.get('price_basis'):
            fails.append('[GATE 8] COMMODITY BASIS: price_basis 未填写')

    # ── v3.0 Gate 1: Archetype purity check ──────────────────────────────────
    if arch == 'commodity':
        if fc.get('price_driver', '').lower() == 'exogenous':
            fails.append('[GATE 1] ARCHETYPE PURITY: commodity archetype 要求 gap-driven ASP，不允许 exogenous')
    if arch == 'project':
        has_penetration = any('penetration' in str(s.get('name', '')).lower() for s in cfg['demand']['segments'])
        if has_penetration:
            warns.append('[GATE 1] ARCHETYPE PURITY: project archetype 不应以 penetration curve 为主需求驱动')
    if arch in ('population', 'installed_base'):
        if fc.get('volume_driver', '').lower().startswith('capacity'):
            fails.append(f'[GATE 1] ARCHETYPE PURITY: {arch} archetype 不能以 capacity 推导需求')

    # ── v3.0 Gate 2: Generator bridge validity for >3Y ───────────────────────
    forecast_yrs = sum(1 for y in cfg.get('years', []) if 'E' in str(y))
    gbv = cfg.get('generator_bridge_validation', {})
    if cfg.get('generator_bridge_required') is True:
        if not cfg.get('demand_engine', {}).get('bridge_nodes'):
            fails.append('[GATE 2] GENERATOR BRIDGE MISSING: bridge_nodes 未填写')
        # Check for pseudo-bridge (no decomposition components)
        if gbv and not any(gbv.values()):
            fails.append('[GATE 2] PSEUDO BRIDGE: generator_bridge_validation 全为 false — 伪 bridge，必须有至少2个分解维度')
        elif gbv:
            true_count = sum(1 for v in gbv.values() if v)
            if true_count < 2:
                warns.append(f'[GATE 2] BRIDGE VALIDITY: generator_bridge_validation 只有 {true_count}/4 维度为 true，建议至少2个')
    elif forecast_yrs > 3:
        warns.append(f'[GATE 2] GENERATOR BRIDGE RECOMMENDED: forecast 有 {forecast_yrs} 个预测年，建议设 generator_bridge_required=true')

    # ── v3.1 Gate 3B: Competition Denominator Consistency ────────────────────
    comp_denom = cfg.get('competition_denominator_basis', {})
    if comp_denom:
        boundary_geo = cfg.get('market_boundary', '')
        comp_geo = comp_denom.get('geography', '')
        if comp_geo and boundary_geo and comp_geo.lower() not in boundary_geo.lower() and boundary_geo.lower() not in comp_geo.lower():
            warns.append(f'[GATE 3B] COMPETITION DENOMINATOR: competition geography={comp_geo} 与 market_boundary 可能不一致，请确认')

    # ── Unit mismatch (legacy, keep as FAIL) ─────────────────────────────────
    for seg in segments:
        seg_vols = _get_seg_vols(seg, yn)
        seg_asps = seg.get('asp_estimates', cfg.get('demand', {}).get('asp_estimates', []))
        if seg_vols and seg_asps:
            vol = seg_vols[mid] if mid < len(seg_vols) else 0
            asp = seg_asps[mid] if mid < len(seg_asps) else 0
            implied = vol * asp
            ru = cfg.get('revenue_unit', '$M')
            scale = 1000 if ru == '$B' else 1
            if implied > 0 and (implied / scale > 1e6 or implied / scale < 1e-3):
                fails.append(f'[UNIT MISMATCH] [{seg["name"]}]: Vol({vol}) × ASP({asp}) = {implied:.0f}, revenue_unit={ru}, ratio={implied/scale:.2e}')

    if not cfg.get('market_boundary'):
        warns.append('MARKET BOUNDARY MISSING: cfg 缺少 market_boundary 字段')

    # ── Write WARNINGS sheet ─────────────────────────────────────────────────
    all_issues = [('FAIL', m) for m in fails] + [('WARN', m) for m in warns]
    if 'WARNINGS' in wb.sheetnames:
        del wb['WARNINGS']
    ws_w = wb.create_sheet('WARNINGS')
    ws_w.column_dimensions['A'].width = 115
    if not all_issues:
        c = ws_w.cell(row=1, column=1, value='✅ All validation checks passed')
        c.font = Font(name='Arial', size=12, bold=True, color='007700')
    else:
        header = '⚠️ MODEL HAS UNRESOLVED ISSUES — DO NOT OUTPUT' if fails else f'Model Validation — 0 FAIL, {len(warns)} WARN'
        color = 'FF0000' if fails else 'CC6600'
        ws_w.cell(row=1, column=1, value=header).font = Font(name='Arial', size=12, bold=True, color=color)
        for i, (level, msg) in enumerate(all_issues, start=2):
            c = ws_w.cell(row=i, column=1, value=f'[{level}] {msg}')
            c.font = Font(name='Arial', size=10, color='CC0000' if level == 'FAIL' else 'CC6600')
    wb.move_sheet('WARNINGS', offset=-len(wb.sheetnames) + 1)

    return fails, warns


if __name__ == '__main__':
    main()
